from openpyxl import Workbook  # 「pip install openpyxl」でインストールしておく

# ワークブックの新規作成と保存
wb = Workbook()
wb.save('myworkbook.xlsx')

# ワークブックの読み込み
from openpyxl import load_workbook

wb = load_workbook('myworkbook.xlsx')

# ワークシートの選択
ws = wb['Sheet']  # ワークシートを指定
ws = wb.active  # アクティブなワークシートを選択
print(f'sheet name: {ws.title}')  # sheet name: Sheet

# ワークシートの作成
wb.create_sheet('my sheet')
wb.create_sheet('my sheet')
wb.create_sheet('my sheet')

# ワークシートの列挙
for sheet in wb:
    print(f'sheet name: {sheet.title}')
# 実行結果：
#Sheet
#my sheet
#my sheet1
#my sheet2

# セルに書き込み
ws['A1'] = 'Hello from Python'
wb.save('myworkbook.xlsx')  # overwrite myworkbook.xlsx

# セルの値の表示
print(ws['a1'].value)  # Hello from Python

# cellメソッドでセルに書き込み
ws.cell(row=1, column=1).value = 1
print(ws['a1'].value)  # 1
ws.cell(row=1, column=1, value=2)
print(ws['a1'].value)  # 2

for idx in range(1, 4):
    ws.cell(row=1, column=idx, value=idx)
    ws.cell(row=2, column=idx, value=-idx)

# ワークシートの全セルを反復（行ごと）
for row in ws.iter_rows():
    for cell in row:
        print(cell.value)
# 出力結果：
#1
#2
#3
#-1
#-2
#-3

for row in ws.rows:
    for cell in row:
        print(cell.value)

min_row = ws.min_row
max_row = ws.max_row
min_col = ws.min_column
max_col = ws.max_column

for row in ws.iter_rows(min_row=min_row, max_row=max_row, 
                        min_col=min_col, max_col=max_col):
    for cell in row:
        print(cell.value)

# セルの値を取得する
for row in ws.iter_rows(values_only=True):
    for value in row:
        print(value)

# ワークシートの全セルを反復（列ごと）
for column in ws.iter_cols(values_only=True):
    print(column)

for column in ws.columns:
    for cell in column:
        print(cell.value)

# セル範囲の取得
rng = ws['A1':'C2']
print(rng)
# 出力結果：
#((<Cell 'Sheet'.A1>, <Cell 'Sheet'.B1>, <Cell 'Sheet'.C1>),
# (<Cell 'Sheet'.A2>, <Cell 'Sheet'.B2>, <Cell 'Sheet'.C2>))

values = []
for row in rng:
    tmp = []
    for col in row:
        tmp.append(col.value)
    values.append(tmp)
#values = [[col.value for col in row] for row in rng]
print(values)

# セル範囲への代入
from openpyxl.cell.cell import Cell

def get_shape(rng):
    return (len(rng), len(rng[0]))

def assign2range(dst, src):
    dst_shape = get_shape(dst)
    src_shape = get_shape(src)
    if src_shape != dst_shape:
        raise ValueError('shapes of arguments not match')

    for d, s in zip(dst, src):  # dst／src: Cellまたは値のタプルを格納するタプル
        for t, v in zip(d, s):  # d／s: Cellまたは値を格納するタプル
            if isinstance(v, Cell):
                v = v.value
            t.value = v

values = [
    [1, 2, 3], [4, 5, 6]
]

assign2range(ws['A1':'C2'], values)
assign2range(ws['A1:C1'], [[100, 200, 300]])
assign2range(ws['A3':'C4'], ws['A1':'C2'])

wb.save('myworkbook.xlsx')