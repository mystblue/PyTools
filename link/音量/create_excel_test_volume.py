# -*- coding: utf-8 -*-

# エビデンスの画像をエクセル形式でまとめる

import openpyxl
from openpyxl import Workbook

device = ["P2 Pro", "P2 Lite SE"]

mode = ["PC 連動", "スタンドアロン"]

rom = ["ROM バージョンアップ前", "ROM バージョンアップ後"]

test_pattern = [
	[
	  "{0}\n　{1}\n　{2}\n　　音量設定",
	  "正常系",
	  "UIテスト",
	  "TMSWEB 管理画面で「ボタンプッシュ音設定」をONに設定する",
	  "音量設定で音量を設定する",
	  "設定した音量になること"
	],
	[
	  "{0}\n　{1}\n　{2}\n　　音量設定",
	  "正常系",
	  "UIテスト",
	  "TMSWEB 管理画面で「ボタンプッシュ音設定」をONに設定する",
	  "音量設定で音量を設定する",
	  "設定した音量になること"
	],
	[
	  "{0}\n　{1}\n　{2}\n　　音量設定",
	  "正常系",
	  "UIテスト",
	  "TMSWEB 管理画面で「ボタンプッシュ音設定」をOFFに設定する",
	  "端末を起動する",
	  "起動時に音量が「0」の状態になること"
	],
	[
	  "{0}\n　{1}\n　{2}\n　　音量設定",
	  "正常系",
	  "UIテスト",
	  "TMSWEB 管理画面で「ボタンプッシュ音設定」をOFFに設定する",
	  "端末を再起動する",
	  "起動時に音量が「0」の状態になること"
	],
]

def do_loop(ws):
    num = 14

    for pattern in test_pattern:
        for d in device:
            for m in mode:
                for r in rom:
                    ws["C" + str(num)] = pattern[0].format(d, m, r)   # テスト項目（評価内容）
                    ws["I" + str(num)] = pattern[1]   # 分類
                    ws["K" + str(num)] = pattern[2]   # テスト観点種別
                    ws["N" + str(num)] = pattern[3]   # 事前条件
                    ws["V" + str(num)] = pattern[4]   # 操作および入力値
                    ws["AL" + str(num)] = pattern[5]   # 振る舞い
                    ws["BI" + str(num)] = "1"
                    num += 1

def report():
    wb = openpyxl.load_workbook("test.xlsx")
    ws = wb['音量機能の修正']  # ワークシートを指定
    ws = wb.active  # アクティブなワークシートを選択
    
    do_loop(ws)

    wb.save('test.xlsx')

if __name__ == '__main__':
    report()
