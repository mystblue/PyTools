# -*- coding: utf-8 -*-

# エビデンスの画像をエクセル形式でまとめる

import openpyxl
from openpyxl import Workbook

device_list = ["Pay TG PC アプリケーション", "Smart TG PC アプリケーション", "Smart TG スタンドアロン"]

check_list = ["磁気・IC利用時のカード預かり登録有効性チェック：空", "磁気・IC利用時のカード預かり登録有効性チェック：1", "磁気・IC利用時のカード預かり登録有効性チェック：0"]

check_list2 = ["カード預かり登録有効性チェック：空", "カード預かり登録有効性チェック：1", "カード預かり登録有効性チェック：0"]

category = ["カードチェック", "オーソリ", "オーソリ[IC]", "オーソリ売上", "オーソリ売上[IC]", "カード預かり登録", "カード預かり更新"]



def do_loop(ws):
    num = 14

    for device in device_list:
        for check in check_list:
            for check2 in check_list2:
                for c in category:
                    title = device + "\n　" + check + "\n　" + check2
                    ws["C" + str(num)] = title
                    ws["I" + str(num)] = "正常系"
                    #ws["K" + str(num)] = # 事前条件
                    ws["S" + str(num)] = c + "を実行する\nカード番号：3580000000001111\n有効期限：05/25" # 操作
                    message = "正常に決済が完了すること\n"
                    if c == 'オーソリ[IC]' or c == 'オーソリ売上[IC]':
                        if check == "磁気・IC利用時のカード預かり登録有効性チェック：0":
                            message += "取引履歴を確認し、有効性チェックなしになっていること（valid_check_flg が 0 になっていること）"
                        else:
                            message += "取引履歴を確認し、有効性チェックありになっていること（valid_check_flg が 1 になっていること）"

                    if c == 'カード預かり登録' or c == 'カード預かり更新':
                        if check2 == "カード預かり登録有効性チェック：0":
                            message += "取引履歴を確認し、有効性チェックなしになっていること（valid_check_flg が 0 になっていること）"
                        else:
                            message += "取引履歴を確認し、有効性チェックありになっていること（valid_check_flg が 1 になっていること）"

                    ws["AI" + str(num)] = message # 振る舞い
                    num += 1

def report():
    wb = openpyxl.load_workbook("test.xlsx")
    ws = wb['決済']  # ワークシートを指定
    ws = wb.active  # アクティブなワークシートを選択
    
    do_loop(ws)

    #ws["C14"] = "test"
    #ws["I14"] = "正常系"
    #ws["Q14"] = "test1"
    #ws["AA14"] = "test2"

    wb.save('test.xlsx')

if __name__ == '__main__':
    report()
