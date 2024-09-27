# -*- coding: utf-8 -*-

# エビデンスの画像をエクセル形式でまとめる

import openpyxl
from openpyxl import Workbook

device = ["PC 連動", "スタンドアロン"]

pc_rendou = [
	[
	  "\n　アプリの自動インストール",
	  "システムテスト",
	  "　　・Smart TG Liteアプリの自動インストールをON\n・Smart TGアプリの自動インストールをON",
	  "アプリがインストールされていない状態で起動する。",
	  "P2 Lite SE\n　・自動でSmart TG Lite アプリがインストールされること。\n　・Smart TGアプリがインストールされないこと。\nP2 Pro\n　・自動でSmart TGアプリがインストールされること。\n　・Smart TG Liteアプリがインストールされないこと。"
	],
	[
	  "\n　アプリの自動インストール",
	  "システムテスト",
	  "　　・Smart TG Liteアプリの自動インストールをOFF\n・Smart TGアプリの自動インストールをON",
	  "アプリがインストールされていない状態で起動する。",
	  "P2 Lite SE\n　・自動でSmart TG Lite アプリがインストールされないこと。\n　・Smart TGアプリがインストールされないこと。"
	],
	[
	  "\n　アプリの自動インストール",
	  "システムテスト",
	  "　　・Smart TG Liteアプリの自動インストールをON\n・Smart TGアプリの自動インストールをOFF",
	  "アプリがインストールされていない状態で起動する。",
	  "P2 Pro\n　・自動でSmart TGアプリがインストールされないこと。\n　・Smart TG Liteアプリがインストールされないこと。"
	],
	[
	  "\n　アプリの自動起動",
	  "システムテスト",
	  "　SUNMIパートナーサイトでSmart TG Liteアプリの自動起動をONにする。\n　SUNMIパートナーサイトでSmart TGアプリの自動起動をONにする。",
	  "決済端末を起動する。",
	  "自動でSmart TG Lite アプリが起動すること。\nSmart TGアプリが起動しないこと。"
	],
	[
	  "\n　アプリの自動起動",
	  "システムテスト",
	  "　SUNMIパートナーサイトでSmart TG Liteアプリの自動起動をOFFにする。\n　SUNMIパートナーサイトでSmart TGアプリの自動起動をONにする。",
	  "決済端末を起動する。",
	  "自動でSmart TG Lite アプリが起動しないこと。\nSmart TGアプリが起動しないこと。"
	],
	[
	  "\n　アプリのバージョンアップ",
	  "システムテスト",
	  "SUNMIパートナーサイトでSmart TG Liteアプリのバージョンアップを実施する。",
	  "決済端末を起動する。",
	  "自動でSmart TG Lite アプリがバージョンアップされること"
	],
	[
	  "\n　アプリのバージョンアップ\n誤ったアプリケーションをアップロードした場合も署名の検証",
	  "システムテスト",
	  "",
	  "SUNMIパートナーサイトでSmart TGアプリのページでSmart TG Liteアプリのapkファイルのアップロードを実施する。",
	  "apkファイルのアップロードが失敗すること。"
	],
	[
	  "\n　アプリのバージョンアップ\n誤ったアプリケーションをアップロードした場合も署名の検証",
	  "システムテスト",
	  "",
	  "SUNMIパートナーサイトでSmart TG LiteアプリのページでSmart TGアプリのapkファイルのアップロードを実施する。",
	  "apkファイルのアップロードが失敗すること。"
	],
]

def do_loop(ws):
    num = 14

    for pattern in pc_rendou:
        for d in device:
            ws["C" + str(num)] = d + pattern[0]   # テスト項目（評価内容）
            ws["I" + str(num)] = "正常系"   # 分類
            ws["K" + str(num)] = pattern[1]   # テスト観点種別
            ws["N" + str(num)] = pattern[2]   # 事前条件
            ws["V" + str(num)] = pattern[3]   # 操作および入力値
            ws["AL" + str(num)] = pattern[4]   # 振る舞い
            ws["BI" + str(num)] = "1"
            num += 1

def report():
    wb = openpyxl.load_workbook("test.xlsx")
    ws = wb['自動インストール']  # ワークシートを指定
    ws = wb.active  # アクティブなワークシートを選択
    
    do_loop(ws)

    wb.save('test.xlsx')

if __name__ == '__main__':
    report()
