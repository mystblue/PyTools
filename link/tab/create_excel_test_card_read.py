# -*- coding: utf-8 -*-

# エビデンスの画像をエクセル形式でまとめる

import openpyxl
from openpyxl import Workbook


devices = ["P2 Pro", "P2 Lite SE"]

modes = ["PC 連動モード", "スタンドアロン"]

methods = ["両方", "磁気/IC読取のみ", "マニュアル入力のみ"]

patterns = ["戻る", "中止"]

displays = ["マニュアル入力タブを表示してクレジットカード番号と有効期限を入力し、磁気/IC読取タブを表示して、IC でカードを読み込み", "磁気/IC読取タブを表示して IC でカードを読み込み、マニュアル入力タブを表示してクレジットカード番号と有効期限を入力し"]

test_cases = [
	[
	  "{mode} {device}\n　両方",
	  "TMSWEB 管理画面で端末の設定からクレジットカード入力方式を「両方」に設定する",
	  "クレジットカードの入力画面を表示し「磁気/IC読取」を選択しクレジットカードを読み取る。",
	  "カード番号と有効期限が読み取れること\nカード番号と有効期限が編集できないこと"
	],
	[
	  "{mode} {device}\n　両方",
	  "TMSWEB 管理画面で端末の設定からクレジットカード入力方式を「両方」に設定する",
	  "クレジットカードの入力画面を表示し「マニュアル入力」を選択しクレジットカードを読み取る。",
	  "カード番号と有効期限が読み取れないこと\nカード番号と有効期限が編集できること"
	],
	[
	  "{mode} {device}\n　磁気/ICのみ",
	  "TMSWEB 管理画面で端末の設定からクレジットカード入力方式を「磁気/ICのみ」に設定する",
	  "クレジットカードの入力画面を表示し、クレジットカードを読み取る。",
	  "カード番号と有効期限が読み取れること\nカード番号と有効期限が編集できないこと"
	],
	[
	  "{mode} {device}\n　マニュアル入力のみ",
	  "TMSWEB 管理画面で端末の設定からクレジットカード入力方式を「マニュアル入力のみ」に設定する",
	  "クレジットカードの入力画面を表示し、クレジットカードを読み取る。",
	  "カード番号と有効期限が読み取れないこと\nカード番号と有効期限が編集できること"
	],
]

test_cases2 = [
	[
	  "タブ切り替え\n{mode} {device}\n　{method}",
	  "TMSWEB 管理画面で端末の設定からクレジットカード入力方式を「{method}」に設定する",
	  "磁気/IC読取からマニュアル入力へタブを切り替える",
	  "カード番号と有効期限クリアされること"
	],
	[
	  "タブ切り替え\n{mode} {device}\n　{method}",
	  "TMSWEB 管理画面で端末の設定からクレジットカード入力方式を「{method}」に設定する",
	  "マニュアル入力から磁気/IC読取へタブを切り替える",
	  "カード番号と有効期限クリアされること"
	],
	[
	  "タブ切り替え\n{mode} {device}\n　{method}",
	  "TMSWEB 管理画面で端末の設定からクレジットカード入力方式を「{method}」に設定する",
	  "磁気/IC読取からマニュアル入力、再度磁気/IC読取へタブを切り替える",
	  "カード番号と有効期限クリアされること"
	],
	[
	  "タブ切り替え\n{mode} {device}\n　{method}",
	  "TMSWEB 管理画面で端末の設定からクレジットカード入力方式を「{method}」に設定する",
	  "磁気/IC読取からマニュアル入力、再度マニュアル入力へタブを切り替える",
	  "カード番号と有効期限クリアされること"
	],
]

test_cases3 = [
	[
	  "{pattern}\n{mode} {device}\n　両方",
	  "TMSWEB 管理画面で端末の設定からクレジットカード入力方式を「両方」に設定する",
	  "クレジットカードの入力画面を表示し、「{pattern}」ボタンを押した後、再度クレジットカード入力画面を表示する",
	  "カード番号と有効期限がクリアされていること\nデフォルト設定されているタブが選択されていること"
	],
	[
	  "{pattern}\n{mode} {device}\n　磁気/IC読取のみ",
	  "TMSWEB 管理画面で端末の設定からクレジットカード入力方式を「磁気/IC読取のみ」に設定する",
	  "クレジットカードの入力画面を表示し、「{pattern}」ボタンを押した後、再度クレジットカード入力画面を表示する",
	  "カード番号と有効期限がクリアされていること"
	],
	[
	  "{pattern}\n{mode} {device}\n　マニュアル入力",
	  "TMSWEB 管理画面で端末の設定からクレジットカード入力方式を「マニュアル入力」に設定する",
	  "クレジットカードの入力画面を表示し、「{pattern}」ボタンを押した後、再度クレジットカード入力画面を表示する",
	  "カード番号と有効期限がクリアされていること"
	],
	[
	  "戻る\n{mode} {device}\n　両方",
	  "TMSWEB 管理画面で端末の設定からクレジットカード入力方式を「両方」に設定する",
	  "クレジットカードの入力画面を表示し、次へを押した後、「{pattern}」ボタンでクレジットカード入力画面に戻る",
	  "カード番号と有効期限が入力されたままになっていること\nデフォルト設定されているタブが選択されていること"
	],
	[
	  "戻る\n{mode} {device}\n　磁気/IC読取のみ",
	  "TMSWEB 管理画面で端末の設定からクレジットカード入力方式を「磁気/IC読取のみ」に設定する",
	  "クレジットカードの入力画面を表示し、次へを押した後、「{pattern}」ボタンでクレジットカード入力画面に戻る",
	  "カード番号と有効期限が入力されたままになっていること"
	],
	[
	  "戻る\n{mode} {device}\n　マニュアル入力",
	  "TMSWEB 管理画面で端末の設定からクレジットカード入力方式を「マニュアル入力」に設定する",
	  "クレジットカードの入力画面を表示し、次へを押した後、「{pattern}」ボタンでクレジットカード入力画面に戻る",
	  "カード番号と有効期限が入力されたままになっていること"
	],
]

test_cases4 = [
	[
	  "どちらのタブでもカード番号を入力して決済を行う\n{mode} {device}\n　両方",
	  "TMSWEB 管理画面で端末の設定からクレジットカード入力方式を「両方」に設定する",
	  "{display}決済を行う",
	  "最後に表示されていたタブで決済できること"
	],
]
def do_loop(ws):
    num = 14

    for pattern in test_cases:
        for d in devices:
            for mode in modes:
                ws["C" + str(num)] = pattern[0].format(mode=mode, device=d)   # テスト項目（評価内容）
                ws["I" + str(num)] = "正常系"   # 分類
                ws["K" + str(num)] = "UIテスト"   # テスト観点種別
                ws["N" + str(num)] = pattern[1]   # 事前条件
                ws["V" + str(num)] = pattern[2]   # 操作及び入力値
                ws["AL" + str(num)] = pattern[3]   # 振る舞い
                ws["BI" + str(num)] = "1" # テスト実施
                num += 1

    for pattern in test_cases2:
        for d in devices:
            for mode in modes:
                for method in methods:
                    ws["C" + str(num)] = pattern[0].format(mode=mode, device=d, method=method)   # テスト項目（評価内容）
                    ws["I" + str(num)] = "正常系"   # 分類
                    ws["K" + str(num)] = "UIテスト"   # テスト観点種別
                    ws["N" + str(num)] = pattern[1].format(method=method)   # 事前条件
                    ws["V" + str(num)] = pattern[2]   # 操作及び入力値
                    ws["AL" + str(num)] = pattern[3]   # 振る舞い
                    ws["BI" + str(num)] = "1" # テスト実施
                    num += 1

    for pattern in test_cases3:
        for d in devices:
            for mode in modes:
                for p in patterns:
                    ws["C" + str(num)] = pattern[0].format(mode=mode, device=d, pattern=p)   # テスト項目（評価内容）
                    ws["I" + str(num)] = "正常系"   # 分類
                    ws["K" + str(num)] = "UIテスト"   # テスト観点種別
                    ws["N" + str(num)] = pattern[1]   # 事前条件
                    ws["V" + str(num)] = pattern[2].format(pattern=p)   # 操作及び入力値
                    ws["AL" + str(num)] = pattern[3]   # 振る舞い
                    ws["BI" + str(num)] = "1" # テスト実施
                    num += 1

    for pattern in test_cases4:
        for d in devices:
            for mode in modes:
                for display in displays:
                    ws["C" + str(num)] = pattern[0].format(mode=mode, device=d)   # テスト項目（評価内容）
                    ws["I" + str(num)] = "正常系"   # 分類
                    ws["K" + str(num)] = "UIテスト"   # テスト観点種別
                    ws["N" + str(num)] = pattern[1]   # 事前条件
                    ws["V" + str(num)] = pattern[2].format(display=display)   # 操作及び入力値
                    ws["AL" + str(num)] = pattern[3]   # 振る舞い
                    ws["BI" + str(num)] = "1" # テスト実施
                    num += 1

def report():
    wb = openpyxl.load_workbook("test.xlsx")
    ws = wb['決済']  # ワークシートを指定
    ws = wb.active  # アクティブなワークシートを選択
    
    do_loop(ws)

    wb.save('test.xlsx')

if __name__ == '__main__':
    report()
