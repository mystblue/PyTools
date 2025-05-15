# -*- coding: utf-8 -*-

# エビデンスの画像をエクセル形式でまとめる

import openpyxl
from openpyxl import Workbook


devices = ["P2 Pro", "P2 Lite SE"]

modes = ["PC 連動モード", "スタンドアロン"]

read_methods = ["両方", "磁気/IC読取", "マニュアル入力"]

result_map = {
  "P2 ProPC 連動モード両方": "フォントサイズ設定\nカード番号入力方式\nバージョン情報",
  "P2 ProPC 連動モード磁気/IC読取": "フォントサイズ設定\nバージョン情報",
  "P2 ProPC 連動モードマニュアル入力": "フォントサイズ設定\nバージョン情報",
  "P2 Proスタンドアロン両方": "決済代行事業者選択\nフォントサイズ設定\nカード番号入力方式\n業務終了\nバージョン情報",
  "P2 Proスタンドアロン磁気/IC読取": "決済代行事業者選択\nフォントサイズ設定\n業務終了\nバージョン情報",
  "P2 Proスタンドアロンマニュアル入力": "決済代行事業者選択\nフォントサイズ設定\n業務終了\nバージョン情報",
  "P2 Lite SEPC 連動モード両方": "フォントサイズ設定\nカード番号入力方式\n音量設定\nバージョン情報",
  "P2 Lite SEPC 連動モード磁気/IC読取": "フォントサイズ設定\n音量設定\nバージョン情報",
  "P2 Lite SEPC 連動モードマニュアル入力": "フォントサイズ設定\n音量設定\nバージョン情報",
  "P2 Lite SEスタンドアロン両方": "決済代行事業者選択\nフォントサイズ設定\nカード番号入力方式\n音量設定\n業務終了\nバージョン情報",
  "P2 Lite SEスタンドアロン磁気/IC読取": "決済代行事業者選択\nフォントサイズ設定\n音量設定\n業務終了\nバージョン情報",
  "P2 Lite SEスタンドアロンマニュアル入力": "決済代行事業者選択\nフォントサイズ設定\n音量設定\n業務終了\nバージョン情報",
}

test_cases = [
	[
	  "{mode} {device}\n　{read_method}",
	  "TMSWEB 管理画面で端末の設定からクレジットカード入力方式を「{read_method}」に設定する",
	  "待受画面で右上の「三」を押してメニューを開く",
	  "メニューに以下が表示されること\n{result}"
	],
]

def do_loop(ws):
    num = 14

    for pattern in test_cases:
        for d in devices:
            for mode in modes:
                for read_method in read_methods:
                    result = result_map[d + mode + read_method]
                    ws["C" + str(num)] = pattern[0].format(mode=mode, device=d, read_method=read_method)   # テスト項目（評価内容）
                    ws["I" + str(num)] = "正常系"   # 分類
                    ws["K" + str(num)] = "UIテスト"   # テスト観点種別
                    ws["N" + str(num)] = pattern[1].format(read_method=read_method)   # 事前条件
                    ws["V" + str(num)] = pattern[2]   # 操作及び入力値
                    ws["AL" + str(num)] = pattern[3].format(result=result)   # 振る舞い
                    ws["BI" + str(num)] = "1" # テスト実施
                    num += 1

def report():
    wb = openpyxl.load_workbook("test.xlsx")
    ws = wb['決済']  # ワークシートを指定
    ws = wb.active  # アクティブなワークシートを選択
    
    do_loop(ws)

    wb.save('test.xlsx')

if __name__ == '__main__':
    report()
