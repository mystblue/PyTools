# -*- coding: utf-8 -*-

# エビデンスの画像をエクセル形式でまとめる

import openpyxl
from openpyxl import Workbook


devices = ["P2 Pro", "P2 Lite SE"]

modes = ["PC 連動モード", "スタンドアロン"]

test_cases = [
	[
	  "{mode} {device}\n　両方",
	  "TMSWEB 管理画面で端末の設定からクレジットカード入力方式を「両方」に設定する",
	  "待受画面のメニューで「カード番号入力方式設定」を選択して、デフォルトで表示されるタブを選択して、「決定」ボタンを押す",
	  "待受画面に戻ること"
	],
	[
	  "{mode} {device}\n　両方",
	  "TMSWEB 管理画面で端末の設定からクレジットカード入力方式を「両方」に設定する",
	  "クレジットカード入力画面を開く",
	  "「カード番号入力方式設定」で選択したタブがデフォルトで表示されること"
	],
]


def do_loop(ws):
    num = 14

    for pattern in test_cases:
        for d in devices:
            for mode in modes:
                ws["C" + str(num)] = pattern[0].format(mode=mode, device=d)   # テスト項目（評価内容）
                ws["I" + str(num)] = "正常系"     # 分類
                ws["K" + str(num)] = "UIテスト"   # テスト観点種別
                ws["N" + str(num)] = pattern[1]   # 事前条件
                ws["V" + str(num)] = pattern[2]   # 操作及び入力値
                ws["AL" + str(num)] = pattern[3]  # 振る舞い
                ws["BI" + str(num)] = "1"         # テスト実施
                num += 1


def report():
    wb = openpyxl.load_workbook("test.xlsx")
    ws = wb['決済']  # ワークシートを指定
    ws = wb.active  # アクティブなワークシートを選択
    
    do_loop(ws)

    wb.save('test.xlsx')

if __name__ == '__main__':
    report()
