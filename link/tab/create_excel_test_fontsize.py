# -*- coding: utf-8 -*-

# エビデンスの画像をエクセル形式でまとめる

import openpyxl
from openpyxl import Workbook


devices = ["P2 Pro", "P2 Lite SE"]

modes = ["PC 連動モード", "スタンドアロン"]

read_methods = ["両方", "磁気/IC読取", "マニュアル入力"]

font_sizes = ["小", "中", "大"]

test_cases = [
	[
	  "{mode} {device}\n　{read_method}",
	  "TMSWEB 管理画面で端末の設定からクレジットカード入力方式を「{read_method}」に設定する",
	  "待受画面のメニューで「フォントサイズ設定」を選択して、フォントサイズを選択して、「決定」ボタンを押すく",
	  "待受画面に戻ること"
	],
]

test_cases2 = [
	[
	  "{mode} {device}",
	  "フォントサイズを{font_size}に設定する",
	  "クレジットカード入力画面で「次へ」を押して取引内容確認画面を表示する",
	  "設定したフォントで表示されること"
	],
	[
	  "{mode} {device}",
	  "フォントサイズを{font_size}に設定する",
	  "決済を実行して、取引結果確認画面を表示する",
	  "設定したフォントで表示されること"
	],
]

def do_loop(ws):
    num = 14

    for pattern in test_cases:
        for d in devices:
            for mode in modes:
                for read_method in read_methods:
                    ws["C" + str(num)] = pattern[0].format(mode=mode, device=d, read_method=read_method)   # テスト項目（評価内容）
                    ws["I" + str(num)] = "正常系"   # 分類
                    ws["K" + str(num)] = "UIテスト"   # テスト観点種別
                    ws["N" + str(num)] = pattern[1].format(read_method=read_method)   # 事前条件
                    ws["V" + str(num)] = pattern[2]   # 操作及び入力値
                    ws["AL" + str(num)] = pattern[3]   # 振る舞い
                    ws["BI" + str(num)] = "1" # テスト実施
                    num += 1

    for pattern in test_cases2:
        for d in devices:
            for mode in modes:
                for font_size in font_sizes:
                    ws["C" + str(num)] = pattern[0].format(mode=mode, device=d)   # テスト項目（評価内容）
                    ws["I" + str(num)] = "正常系"     # 分類
                    ws["K" + str(num)] = "UIテスト"   # テスト観点種別
                    ws["N" + str(num)] = pattern[1].format(font_size=font_size)   # 事前条件
                    ws["V" + str(num)] = pattern[2]   # 操作及び入力値
                    ws["AL" + str(num)] = pattern[3]  # 振る舞い
                    ws["BI" + str(num)] = "1"         # テスト実施
                    num += 1

def report():
    wb = openpyxl.load_workbook("test.xlsx")
    ws = wb['決済']  # ワークシートを指定
    ws = wb.active  # アクティブなワークシートを選択
    
    do_loop(ws)

    wb.save('test.xlsx')

if __name__ == '__main__':
    report()
