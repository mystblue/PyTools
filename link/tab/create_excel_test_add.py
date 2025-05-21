# -*- coding: utf-8 -*-

# エビデンスの画像をエクセル形式でまとめる

import openpyxl
from openpyxl import Workbook


devices = ["P2 Pro", "P2 Lite SE"]

modes = ["PC 連動モード", "スタンドアロン"]

test_cases = [
	[
	  "{mode} {device}\n　設定同期",
	  "フォントサイズ大に設定してある端末を用意する",
	  "フォントサイズ大に設定してあるシリアル番号の端末設定を設定同期対象として、設定同期を行う",
	  "フォントサイズの設定は引き継がれないこと"
	],
	[
	  "{mode} {device}\n　再起動",
	  "フォントサイズ大に設定する",
	  "端末を再起動する",
	  "フォントサイズの設定を開き、大のままになっていること"
	],
	[
	  "{mode} {device}\n　設定同期",
	  "カード番号入力方式でマニュアル入力に設定してある端末を用意する",
	  "カード番号入力方式でマニュアル入力に設定してある端末を設定同期対象として、設定同期を行う",
	  "カード番号入力方式で設定したデフォルトタブは設定同期の対象となっておらず、引き継がれないこと"
	],
	[
	  "{mode} {device}\n　再起動",
	  "カード番号入力方式でマニュアル入力に設定する",
	  "端末を再起動する",
	  "カード番号入力方式の設定を開き、マニュアル入力のままになっていること"
	],
]

def do_loop(ws):
    num = 14

    for pattern in test_cases:
        for d in devices:
            for mode in modes:
                ws["C" + str(num)] = pattern[0].format(mode=mode, device=d)   # テスト項目（評価内容）
                ws["I" + str(num)] = "正常系"   # 分類
                ws["K" + str(num)] = "UIテスト"   # テスト観点種別
                ws["N" + str(num)] = pattern[1]   # 事前条件
                ws["V" + str(num)] = pattern[2]   # 操作及び入力値
                ws["AL" + str(num)] = pattern[3]   # 振る舞い
                ws["BI" + str(num)] = "1" # テスト実施
                num += 1

def report():
    wb = openpyxl.load_workbook("test.xlsx")
    ws = wb['決済']  # ワークシートを指定
    ws = wb.active  # アクティブなワークシートを選択
    
    do_loop(ws)

    wb.save('test.xlsx')

if __name__ == '__main__':
    report()
