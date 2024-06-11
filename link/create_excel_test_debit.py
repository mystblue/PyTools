# -*- coding: utf-8 -*-

# エビデンスの画像をエクセル形式でまとめる

import openpyxl
from openpyxl import Workbook

device_list = ["デビプリ判定フラグ：未入力", "デビプリ判定フラグ：0", "デビプリ判定フラグ：1"]

check_list = ["オーソリ", "オーソリ売上", "カードチェック", "カード預かり登録", "カード預かり更新"]

input_methods = [("クレジットカード", "カード番号：4111111111111111"), ("デビットカード", "カード番号：4001420000000000"), ("プリペイドカード", "カード番号：4063190000000000")]


content = "全パラメータをバリデーションルールに従って入力し、決済が正常に終了すること"

message = "取引履歴に記録されていることを確認する"

def do_loop(ws):
    num = 14

    for device in device_list:
        for check in check_list:
            for input_method in input_methods:
                title = "ペイジェント デビット／プリペイド判定\n" + device + "\n" + "　" + check + "\n　" + input_method[0]
                ws["C" + str(num)] = title
                ws["I" + str(num)] = "正常系"
                ws["S" + str(num)] = check + "\n" + input_method[1]
                if device == "デビプリ判定フラグ：1" and (check == "カード預かり登録" or check == "カード預かり登録") and input_method[0] != "クレジットカード":
                    ws["AI" + str(num)] = "GW エラーとなり、「入力されたカード番号はデビットカード/プリペイドカードのため、ご利用いただくことができません。クレジットカードのご利用をお願いします。」というメッセージが表示されること"
                else:
                    ws["AI" + str(num)] = "正常に決済できること"
                num += 1

def report():
    wb = openpyxl.load_workbook("test.xlsx")
    ws = wb['決済']  # ワークシートを指定
    ws = wb.active  # アクティブなワークシートを選択
    
    do_loop(ws)

    #ws["C14"] = "test"
    #ws["I14"] = "正常系"
    #ws["Q14"] = "test1"
    #ws["AA14"] = "test2"

    wb.save('test.xlsx')

if __name__ == '__main__':
    report()
