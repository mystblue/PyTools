# -*- coding: utf-8 -*-

# エビデンスの画像をエクセル形式でまとめる

import openpyxl
from openpyxl import Workbook

test_pattern = [
	[
	  "TMSWEB\n　端末画面",
	  "正常系",
	  "UIテスト",
	  "端末一覧画面を開く",
	  "「新規登録」ボタンを押す",
	  "「LTE 自動切換(Pay TG) / Wi-Fi 許可(Smart TG)」が表示されること"
	],
	[
	  "TMSWEB\n　端末画面",
	  "正常系",
	  "UIテスト",
	  "端末一覧画面を開く",
	  "「詳細」ボタンを押す",
	  "「LTE 自動切換(Pay TG) / Wi-Fi 許可(Smart TG)」が表示されること"
	],
	[
	  "TMSWEB\n　端末画面",
	  "正常系",
	  "UIテスト",
	  "端末一覧画面を開く",
	  "「編集」ボタンを押す",
	  "「LTE 自動切換(Pay TG) / Wi-Fi 許可(Smart TG)」が表示されること"
	],
	[
	  "TMSWEB\n　CSVダウンロード",
	  "正常系",
	  "UIテスト",
	  "端末一覧画面を開く",
	  "「ダウンロード」ボタンを押す",
	  "ダウンロードしたCSVのヘッダが「LTE 自動切換(Pay TG) / Wi-Fi 許可(Smart TG)」となっていること"
	],
	[
	  "TMSWEB\n　端末一括登録",
	  "正常系",
	  "UIテスト",
	  "端末一覧画面を開く",
	  "ヘッダが「LTE 自動切換(Pay TG) / Wi-Fi 許可(Smart TG)」となっている CSV を一括登録する",
	  "正常に一括登録できること"
	],
	[
	  "TMSWEB\n　端末一括登録",
	  "正常系",
	  "UIテスト",
	  "端末一覧画面を開く",
	  "ヘッダが「LTE 自動切換」となっている CSV を一括登録する",
	  "正常に一括登録できること"
	],
]

def do_loop(ws):
    num = 14

    for pattern in test_pattern:
        ws["C" + str(num)] = pattern[0]   # テスト項目（評価内容）
        ws["I" + str(num)] = pattern[1]   # 分類
        ws["K" + str(num)] = pattern[2]   # テスト観点種別
        ws["N" + str(num)] = pattern[3]   # 事前条件
        ws["V" + str(num)] = pattern[4]   # 操作および入力値
        ws["AL" + str(num)] = pattern[5]   # 振る舞い
        ws["BI" + str(num)] = "1"
        num += 1

def report():
    wb = openpyxl.load_workbook("test.xlsx")
    ws = wb['TMSWEB']  # ワークシートを指定
    ws = wb.active  # アクティブなワークシートを選択
    
    do_loop(ws)

    wb.save('test.xlsx')

if __name__ == '__main__':
    report()
