# -*- coding: utf-8 -*-

# エビデンスの画像をエクセル形式でまとめる

import openpyxl
from openpyxl import Workbook

device = ["P2 Pro", "P2 Lite SE"]

mode = ["PC 連動", "スタンドアロン"]

rom = ["ROM バージョンアップ前", "ROM バージョンアップ後"]

test_pattern = [
	[
	  "{0}\n　{1}\n　{2}\n　　画面設定",
	  "正常系",
	  "UIテスト",
	  "",
	  "端末を起動し、ホーム画面からメニューを開く",
	  "メニューに「画面設定」が表示されること"
	],
	[
	  "{0}\n　{1}\n　{2}\n　　画面設定",
	  "正常系",
	  "UIテスト",
	  "",
	  "ホーム画面のメニューから「画面設定」を押下する",
	  "画面設定画面に切り替わること"
	],
	[
	  "{0}\n　{1}\n　{2}\n　　輝度設定",
	  "正常系",
	  "UIテスト",
	  "",
	  "「画面設定」で画面の輝度を変更する",
	  "設定した値に応じて、画面の輝度が変更されること\nまた実際の画面の輝度が設定した明るさに変わること"
	],
]

test_pattern3 = [
	[
	  "{0}\n　{1}\n　{2}\n　　デフォルト値の確認",
	  "正常系",
	  "機能テスト",
	  "今回リリースするバージョンにアップデートする",
	  "メニューから画面設定を開く",
	  "輝度がデフォルト値になっていること\n実際の画面の輝度がデフォルト値の明るさになっていること"
	],
]

def do_loop(ws):
    num = 14

    for pattern in test_pattern:
        for d in device:
            for m in mode:
                for r in rom:
                    ws["C" + str(num)] = pattern[0].format(d, m, r)   # テスト項目（評価内容）
                    ws["I" + str(num)] = pattern[1]   # 分類
                    ws["K" + str(num)] = pattern[2]   # テスト観点種別
                    ws["N" + str(num)] = pattern[3]   # 事前条件
                    ws["V" + str(num)] = pattern[4]   # 操作および入力値
                    ws["AL" + str(num)] = pattern[5]   # 振る舞い
                    ws["BI" + str(num)] = "1"
                    num += 1

    for pattern in test_pattern3:
        for d in device:
            for m in mode:
                for r in rom:
                    ws["C" + str(num)] = pattern[0].format(d, m, r)   # テスト項目（評価内容）
                    ws["I" + str(num)] = pattern[1]   # 分類
                    ws["K" + str(num)] = pattern[2]   # テスト観点種別
                    ws["N" + str(num)] = pattern[3]   # 事前条件
                    ws["V" + str(num)] = pattern[4]   # 操作および入力値
                    ws["AL" + str(num)] = pattern[5]   # 振る舞い
                    ws["BI" + str(num)] = "1"
                    num += 1

def report():
    wb = openpyxl.load_workbook("test.xlsx")
    ws = wb['バッテリー膨張対策']  # ワークシートを指定
    ws = wb.active  # アクティブなワークシートを選択
    
    do_loop(ws)

    wb.save('test.xlsx')

if __name__ == '__main__':
    report()
