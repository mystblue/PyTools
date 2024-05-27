# -*- coding: utf-8 -*-

# エビデンスの画像をエクセル形式でまとめる

import openpyxl
from openpyxl import Workbook

psp_list = ["DGFT", "ソニーペイメント", "ペイジェント", "GMO-PG", "SMBCファイナンス", "ヤマトフィナンシャル", "ゼウス", "SMBC-GMO", "GMO-MG", "NTTコムオンライン", "楽天カード", "pay.jp", "sbps"]

input_methods = [("同期端末が存在しない", "設定同期を行い、同期する端末が存在しないメッセージが表示されること"), ("同期端末が一つだけ", "設定同期を行い、同期端の設定が反映されること"), ("同期端末が複数存在", "設定同期を行い、同期する端末を選択し、選択した端末の設定が反映されること")]


content = "全パラメータをバリデーションルールに従って入力し、決済が正常に終了すること"

message = "取引履歴に記録されていることを確認する"

def do_loop(ws):
    num = 14

    for psp in psp_list:
        for input_method in input_methods:
            title = "設定PSP：" + psp
            ws["C" + str(num)] = title
            ws["I" + str(num)] = "正常系"
            ws["K" + str(num)] = input_method[0]
            ws["S" + str(num)] = input_method[1]
            num += 1

def report():
    wb = openpyxl.load_workbook("test.xlsx")
    ws = wb['決済']  # ワークシートを指定
    ws = wb.active  # アクティブなワークシートを選択
    
    do_loop(ws)

    #ws["C14"] = "test"
    #ws["I14"] = "正常系"
    #ws["Q14"] = "test1"
    #ws["AA14"] = "test2"

    wb.save('test.xlsx')

if __name__ == '__main__':
    report()
