# -*- coding: utf-8 -*-

# エビデンスの画像をエクセル形式でまとめる

import openpyxl
from openpyxl import Workbook

psp = [
    "DGFT",
    "ソニーペイメントサービス",
    "ペイジェント",
    "GMO-PG",
    "SMBCファイナンス",
    "ヤマト運輸",
    "ゼウス",
    "SMBC-GMO",
    "GMO-MB",
    "NTTコムオンライン",
    "楽天カード",
    "PAY.JP",
    "ソフトバンクペイメントサービス"
]

pw_list = [
    "ソニーペイメントサービス",
    "ペイジェント",
    "GMO-PG",
    "SMBCファイナンス",
    "SMBC-GMO",
]

fonts = ["スタンドアロン（フォントサイズ：小）", "スタンドアロン（フォントサイズ：中）", "スタンドアロン（フォントサイズ：大）"]

pc_rendou = [
	[
	  "PC 連動モード\n　待受画面",
	  "UIテスト",
	  "",
	  "待受画面を開く",
	  "レイアウトが崩れていないこと"
	],
	[
	  "PC 連動モード\n　電源OFFを押した時の終了画面",
	  "UIテスト",
	  "",
	  "電源OFFボタンをタップして終了確認画面を開く",
	  "レイアウトが崩れていないこと"
	],
	[
	  "PC 連動モード\n　カード番号入力画面",
	  "UIテスト",
	  "クレジットカード読取方式を両方に設定する",
	  "PC アプリで「実行」ボタンをクリックしてカード番号入力画面を開く",
	  "レイアウトが崩れていないこと"
	],
	[
	  "PC 連動モード\n　カード番号入力画面",
	  "UIテスト",
	  "クレジットカード読取方式を磁気／IC読取のみに設定する",
	  "PC アプリで「実行」ボタンをクリックしてカード番号入力画面を開く",
	  "レイアウトが崩れていないこと"
	],
	[
	  "PC 連動モード\n　カード番号入力画面",
	  "UIテスト",
	  "クレジットカード読取方式をマニュアル入力のみに設定する",
	  "PC アプリで「実行」ボタンをクリックしてカード番号入力画面を開く",
	  "レイアウトが崩れていないこと"
	],
	[
	  "PC 連動モード\n　セキュリティコード入力画面",
	  "UIテスト",
	  "セキュリティコード入力設定を「ON」にする",
	  "クレジットカード番号入力画面でカード番号と有効期限を入力し、「次へ」をタップする",
	  "レイアウトが崩れていないこと"
	],
	[
	  "PC 連動モード\n　決済完了画面",
	  "UIテスト",
	  "",
	  "クレジットカード番号を入力して決済を完了させる",
	  "レイアウトが崩れていないこと"
	],
	[
	  "PC 連動モード\n　通信失敗画面",
	  "UIテスト",
	  "GW のインスタンを終了させる",
	  "決済を行う",
	  "GW 接続エラーが表示されること\nレイアウトが崩れていないこと"
	],
]

standalone = [
    [
      "\n　決済代行事業者選択画面",
	  "UIテスト",
      "シングルの加盟店に端末を紐付ける",
      "…メニューから「決済代行事業者選択」をタップする",
      "レイアウトが崩れていないこと"
    ],
    [
      "\n　決済代行事業者選択画面",
	  "UIテスト",
      "マルチPSPの加盟店に端末を紐付ける",
      "…メニューから「決済代行事業者選択」をタップする",
      "レイアウトが崩れていないこと\nスクロールしてもレイアウトが崩れないこと"
    ],
    [
      "\n　設定同期画面",
	  "UIテスト",
      "設定同期端末なし",
      "決済代行事業者選択画面で設定同期をタップする",
      "設定同期する端末がないメッセージがダイアログで表示されること\nレイアウトが崩れないこと"
    ],
    [
      "\n　設定同期画面",
	  "UIテスト",
      "設定同期端末１台のみ",
      "決済代行事業者選択画面で設定同期をタップする",
      "シリアル番号選択画面が表示されること\nレイアウトが崩れないこと"
    ],
    [
      "\n　設定同期画面",
	  "UIテスト",
      "設定同期端末が１０台以上",
      "決済代行事業者選択画面で設定同期をタップする",
      "シリアル番号選択画面が表示されること\nレイアウトが崩れないこと\nスクロールしてもレイアウトが崩れないこと"
    ],
    [
      "\n　設定同期画面\n　シリアル番号確認ダイアログ",
	  "UIテスト",
      "",
      "シリアル番号選択画面でシリアル番号をタップする",
      "シリアル番号選択確認ダイアログが表示されること\nレイアウトが崩れないこと"
    ],
    [
      "\n　設定同期画面\n　終了ダイアログ",
	  "UIテスト",
      "",
      "シリアル番号確認ダイアログで「はい」をタップする",
      "「設定同期を終了します」ダイアログが表示されること\nレイアウトが崩れないこと"
    ],
]

standalone2 = [
    [
      "\n　決済代行事業者設定登録\n　決済代行事業者発行ID/PW登録画面\n　　　",
	  "UIテスト",
      "マルチPSPモードに設定",
      "取引種別画面で「次へ」をタップする",
      "決済代行事業者発行ID/PW登録画面が表示されること\nレイアウトが崩れていないこと"
    ],
    [
      "\n　決済代行事業者設定登録\n　取引種別画面\n　　　",
	  "UIテスト",
      "",
      "「登録／編集」をタップする",
      "取引種別選択画面が開くこと\nレイアウトが崩れていないこと"
    ],
    [
      "\n　決済代行事業者設定登録\n　決済取引用パラメータ設定画面\n　　　",
	  "UIテスト",
      "",
      "取引種別画面で「次へ」をタップする",
      "決済取引用パラメータ設定画面が表示されること\nレイアウトが崩れていないこと"
    ],
    [
      "\n　決済代行事業者設定登録\n　決済取引用パラメータ設定画面\n　　　パラメータ展開\n　　　　",
	  "UIテスト",
      "",
      "決済取引用パラメータ設定画面で取引種別をタップして各パラメータを表示させる",
      "全パラメータを表示してレイアウトが崩れないこと\nスクロールしてもレイアウトが崩れていないこと"
    ],
    [
      "\n　決済代行事業者設定登録\n　決済取引用パラメータ確認画面\n　　　パラメータ展開\n　　　　",
	  "UIテスト",
      "",
      "決済取引用パラメータ設定画面で取引種別をタップして各パラメータを表示させる",
      "全パラメータを表示してレイアウトが崩れないこと\nスクロールしてもレイアウトが崩れていないこと"
    ],
    [
      "\n　決済データ入力\n　",
	  "UIテスト",
      "",
      "端末を起動して決済データ入力画面を表示する",
      "レイアウトが崩れないこと\nドロップダウンリストをタップしても項目が正しく表示されること"
    ],
    [
      "\n　決済データ入力\n　パラメータ入力画面\n　　　",
	  "UIテスト",
      "",
      "決済データ入力画面で「決済データ入力」をタップする",
      "決済データ入力画面が開くこと\nレイアウトが崩れないこと\nスクロールしてもレイアウトが崩れないこと"
    ],
    [
      "\n　決済データ入力\n　クレジットカード情報入力画面\n　　　",
	  "UIテスト",
	  "クレジットカード読取方式を両方に設定する",
      "決済データ入力画面で「次へ」をタップする",
      "レイアウトが崩れないこと"
    ],
    [
      "\n　決済データ入力\n　クレジットカード情報入力画面\n　　　",
	  "UIテスト",
	  "クレジットカード読取方式を磁気／IC読取のみに設定する",
      "決済データ入力画面で「次へ」をタップする",
      "レイアウトが崩れないこと"
    ],
    [
      "\n　決済データ入力\n　クレジットカード情報入力画面\n　　　",
	  "UIテスト",
	  "クレジットカード読取方式をマニュアル入力のみに設定する",
      "決済データ入力画面で「次へ」をタップする",
      "レイアウトが崩れないこと"
    ],
    [
      "\n　決済データ入力\n　セキュリティコード入力画面\n　　　",
	  "UIテスト",
	  "セキュリティコード入力設定を「ON」にする",
      "クレジットカード情報入力画面で「次へ」をタップする",
      "レイアウトが崩れないこと"
    ],
    [
      "\n　決済データ入力\n　取引内容確認画面\n　　　",
	  "UIテスト",
	  "",
      "セキュリティコード入力画面で「次へ」をタップする",
      "レイアウトが崩れないこと\nスクロールしてもレイアウトが崩れないこと"
    ],
    [
      "\n　決済データ入力\n　処理結果画面画面\n　　　",
	  "UIテスト",
	  "",
      "決済を実行する",
      "レイアウトが崩れないこと\nスクロールしてもレイアウトが崩れないこと"
    ],
]

standalone3 = [
	[
	  "\n　通信失敗画面",
	  "UIテスト",
	  "GW のインスタンを終了させる",
	  "決済を行う",
	  "GW 接続エラーが表示されること\nレイアウトが崩れていないこと"
	],
    [
      "\n　業務終了画面",
	  "UIテスト",
      "",
      "…メニューから「フォントサイズ設定」をタップする",
      "フォントサイズ設定画面が開くこと\nレイアウトが崩れていないこと"
    ],
    [
      "\n　バージョン情報",
	  "UIテスト",
      "",
      "…メニューから「バージョン情報」をタップする",
      "バージョン情報が表示されること\nレイアウトが崩れていないこと"
    ],
]

def do_loop(ws):
    num = 14
    
    for pattern in pc_rendou:
        ws["C" + str(num)] = pattern[0]   # テスト項目（評価内容）
        ws["I" + str(num)] = "正常系"   # 分類
        ws["K" + str(num)] = pattern[1]   # テスト観点種別
        ws["N" + str(num)] = pattern[2]   # 事前条件
        ws["V" + str(num)] = pattern[3]   # 操作および入力値
        ws["AL" + str(num)] = pattern[4]   # 振る舞い
        ws["BI" + str(num)] = "1"
        num += 1
    
    for pattern in standalone:
        for font in fonts:
            ws["C" + str(num)] = font + pattern[0]   # テスト項目（評価内容）
            ws["I" + str(num)] = "正常系"   # 分類
            ws["K" + str(num)] = pattern[1]   # テスト観点種別
            ws["N" + str(num)] = pattern[2]   # 事前条件
            ws["V" + str(num)] = pattern[3]   # 操作および入力値
            ws["AL" + str(num)] = pattern[4]   # 振る舞い
            ws["BI" + str(num)] = "1"
            num += 1
    
    for p in psp:
        for pattern in standalone2:
            for font in fonts:
                if pattern[2] == "マルチPSPモードに設定":
                    ws["C" + str(num)] = font + pattern[0] + p   # テスト項目（評価内容）
                    ws["I" + str(num)] = "正常系"   # 分類
                    ws["K" + str(num)] = pattern[1]   # テスト観点種別
                    ws["N" + str(num)] = "シングルモードに設定"   # 事前条件
                    ws["V" + str(num)] = pattern[3]   # 操作および入力値
                    ws["AL" + str(num)] = pattern[4]   # 振る舞い
                    ws["BI" + str(num)] = "1"
                    num += 1
                ws["C" + str(num)] = font + pattern[0] + p   # テスト項目（評価内容）
                ws["I" + str(num)] = "正常系"   # 分類
                ws["K" + str(num)] = pattern[1]   # テスト観点種別
                ws["N" + str(num)] = pattern[2]   # 事前条件
                ws["V" + str(num)] = pattern[3]   # 操作および入力値
                ws["AL" + str(num)] = pattern[4]   # 振る舞い
                ws["BI" + str(num)] = "1"
                num += 1
    
    for pattern in standalone3:
        for font in fonts:
            ws["C" + str(num)] = font + pattern[0]   # テスト項目（評価内容）
            ws["I" + str(num)] = "正常系"   # 分類
            ws["K" + str(num)] = pattern[1]   # テスト観点種別
            ws["N" + str(num)] = pattern[2]   # 事前条件
            ws["V" + str(num)] = pattern[3]   # 操作および入力値
            ws["AL" + str(num)] = pattern[4]   # 振る舞い
            ws["BI" + str(num)] = "1"
            num += 1

def report():
    wb = openpyxl.load_workbook("test.xlsx")
    ws = wb['画面サイズの変更']  # ワークシートを指定
    ws = wb.active  # アクティブなワークシートを選択
    
    do_loop(ws)
    
    wb.save('test.xlsx')

if __name__ == '__main__':
    report()