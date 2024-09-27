# -*- coding: utf-8 -*-

# エビデンスの画像をエクセル形式でまとめる

import openpyxl
from openpyxl import Workbook
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment
from openpyxl.worksheet.datavalidation import DataValidation

device = ["P2 Lite SE PC 連動", "P2 Lite SE スタンドアロン"]

pc_rendou = [
    # 公開鍵の取得
	[
	  "\n　公開鍵の取得",
	  "機能テスト\nログ・モニタリングテスト",
	  "公開鍵(バージョン3000)がSecrets Managerに登録されている",
	  "決済端末にSmart TGアプリをインストールし起動する。",
	  "公開鍵の取得ログが記録されること\n決済端末の初期画面が表示されること。\nコンテナ、CloudTail に取得ログが出力されること"
	],
    # 公開鍵の取得失敗
	[
	  "\n　公開鍵の取得に失敗",
	  "エラーハンドリングテスト\nログ・モニタリングテスト",
	  "端末のシリアル番号が存在しない",
	  "決済端末にSmart TGアプリをインストールし起動する。",
	  "公開鍵の取得ログが記録されないこと\nコンテナ、CloudTail に取得ログが出力されないこと"
	],
	[
	  "\n　公開鍵の取得に失敗",
	  "エラーハンドリングテスト\nログ・モニタリングテスト",
	  "端末が無効化されている",
	  "決済端末にSmart TGアプリをインストールし起動する。",
	  "公開鍵の取得ログが記録されないこと\nコンテナ、CloudTail に取得ログが出力されないこと"
	],
	[
	  "\n　公開鍵の取得に失敗",
	  "エラーハンドリングテスト\nログ・モニタリングテスト",
	  "該当の公開鍵がSecrets Managerに存在しない",
	  "決済端末にSmart TGアプリをインストールし起動する。",
	  "公開鍵の取得ログが記録されないこと\nコンテナ、CloudTail に取得ログが出力されないこと"
	],
    # 公開鍵の誤り
	[
	  "\n　公開鍵の誤り",
	  "エラーハンドリングテスト\nログ・モニタリングテスト",
	  "公開鍵(バージョン3000)に誤りがある",
	  "決済端末にSmart TGアプリをインストールし起動する。",
	  "公開鍵の取得ログが記録されること\nコンテナ、CloudTail に取得ログが出力されること"
	],
    # TMS Seed の取得
	[
	  "\n　TMS との通信",
	  "機能テスト\nログ・モニタリングテスト",
	  "TMS Seed(バージョン3000)がSecrets Managerに登録されている。",
	  "決済端末にSmart TGアプリをインストールし起動する。",
	  "アプリケーションセット>アップデートログに履歴が記録されること。\n決済端末の初期画面が表示されること。\nコンテナ、CloudTail に取得ログが出力されること"
	],
    # TMS Seed の取得失敗
	[
	  "\n　TMS との通信失敗",
	  "エラーハンドリングテスト\nログ・モニタリングテスト",
	  "端末のシリアル番号が存在しない",
	  "決済端末にSmart TGアプリをインストールし起動する。",
	  "TMS Seed(バージョン3000)の取得ログが記録されないこと。\nコンテナ、CloudTail に取得ログが出力されないこと"
	],
	[
	  "\n　TMS との通信失敗",
	  "エラーハンドリングテスト\nログ・モニタリングテスト",
	  "端末が無効化されている",
	  "決済端末にSmart TGアプリをインストールし起動する。",
	  "TMS Seed(バージョン3000)の取得ログが記録されないこと。\nコンテナ、CloudTail に取得ログが出力されないこと"
	],
	[
	  "\n　TMS との通信失敗",
	  "エラーハンドリングテスト\nログ・モニタリングテスト",
	  "該当のTMS SeedがSecrets Managerに存在しない",
	  "決済端末にSmart TGアプリをインストールし起動する。",
	  "TMS Seed(バージョン3000)の取得ログが記録されないこと。\nコンテナ、CloudTail に取得ログが出力されないこと"
	],
    # TMS Seed の誤り
	[
	  "\n　TMS Seed の誤り",
	  "エラーハンドリングテスト\nログ・モニタリングテスト",
	  "TMS Seed(バージョン3000)に誤りがある",
	  "決済端末にSmart TGアプリをインストールし起動する。",
	  "TMS Seed 取得ログが記録されること\nコンテナ、CloudTail に取得ログが出力されること"
	],
    # GW Seed の取得
	[
	  "\n　GW との通信",
	  "機能テスト\nログ・モニタリングテスト",
	  "GW Seed(バージョン3000)がSecrets Managerに登録されている。",
	  "決済端末にSmart TGアプリをインストールし起動する。",
	  "アプリケーションセット>アップデートログに履歴が記録されること。決済端末の初期画面が表示されること。\nコンテナ、CloudTail に取得ログが出力されること"
	],
    # GW Seed の取得失敗
	[
	  "\n　GW との通信失敗",
	  "エラーハンドリングテスト\nログ・モニタリングテスト",
	  "端末のシリアル番号が存在しない",
	  "決済端末にSmart TGアプリをインストールし起動する。",
	  "GW Seed(バージョン3000)の取得ログが記録されないこと。\nコンテナ、CloudTail に取得ログが出力されないこと"
	],
	[
	  "\n　GW との通信失敗",
	  "エラーハンドリングテスト\nログ・モニタリングテスト",
	  "端末が無効化されている",
	  "決済端末にSmart TGアプリをインストールし起動する。",
	  "GW Seed(バージョン3000)の取得ログが記録されないこと。\nコンテナ、CloudTail に取得ログが出力されないこと"
	],
	[
	  "\n　GW との通信失敗",
	  "エラーハンドリングテスト\nログ・モニタリングテスト",
	  "該当のGW SeedがSecrets Managerに存在しない",
	  "決済端末にSmart TGアプリをインストールし起動する。",
	  "GW Seed(バージョン3000)の取得ログが記録されないこと。\nコンテナ、CloudTail に取得ログが出力されないこと"
	],
    # GW Seed の誤り
	[
	  "\n　GW Seed の誤り",
	  "エラーハンドリングテスト\nログ・モニタリングテスト",
	  "GW Seed(バージョン3000)に誤りがある",
	  "決済端末にSmart TGアプリをインストールし起動する。",
	  "GW Seed 取得ログが記録されること\nコンテナ、CloudTail に取得ログが出力されること"
	],
]

side = Side(style='thin', color='000000')
border = Border(top=side, bottom=side, left=side, right=side)

def do_loop(ws):
    num = 14

    dv = DataValidation(
        type="list",
        formula1='"正常系,異常系,準正常系"',
        allow_blank=True,
        showErrorMessage=False,
    )


    ws['A' + str(num)].border = border
    ws['B' + str(num)].border = border

    ws.merge_cells('A' + str(num) + ':B' + str(num))
    
    ws['A' + str(num)] = '=ROW()-13'
    
    # https://gammasoft.jp/blog/text-center-alignment-with-openpyxl/#method-2
    ws['A' + str(num)].alignment = Alignment(horizontal='center', vertical='center')
    
    ws.row_dimensions[num].height = 90



    ws['C' + str(num)].border = border
    ws['D' + str(num)].border = border
    ws['E' + str(num)].border = border
    ws['F' + str(num)].border = border
    ws['G' + str(num)].border = border
    ws['H' + str(num)].border = border

    ws.merge_cells('C' + str(num) + ':H' + str(num))
    
    ws['C' + str(num)] = 'GW Seed の誤り'
    
    # https://gammasoft.jp/blog/text-center-alignment-with-openpyxl/#method-2
    ws['C' + str(num)].alignment = Alignment(horizontal='left', vertical='top', wrapText=True)
    


    ws['I' + str(num)].border = border
    ws['J' + str(num)].border = border

    ws.merge_cells('I' + str(num) + ':J' + str(num))
    
    dv.add(f'I' + str(num))
    #ws['I' + str(num)] = '正常系'
    
    # https://gammasoft.jp/blog/text-center-alignment-with-openpyxl/#method-2
    ws['I' + str(num)].alignment = Alignment(horizontal='center', vertical='center')

    ws.add_data_validation(dv)

"""
    ws["C" + str(num)] = d + pattern[0]   # テスト項目（評価内容）
    ws["I" + str(num)] = "正常系"   # 分類
    ws["K" + str(num)] = pattern[1]   # テスト観点種別
    ws["N" + str(num)] = pattern[2]   # 事前条件
    ws["V" + str(num)] = pattern[3]   # 操作および入力値
    ws["AL" + str(num)] = pattern[4]   # 振る舞い
    ws["BI" + str(num)] = "1"
"""
    #num += 1

def report():
    wb = openpyxl.load_workbook("test.xlsx")
    ws = wb['公開鍵・Seed']  # ワークシートを指定
    ws = wb.active  # アクティブなワークシートを選択
    
    do_loop(ws)

    wb.save('test.xlsx')

if __name__ == '__main__':
    report()
