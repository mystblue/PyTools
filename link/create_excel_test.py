# -*- coding: utf-8 -*-

# エビデンスの画像をエクセル形式でまとめる

import openpyxl
from openpyxl import Workbook

psp_list = ["DGFT", "ソニーペイメント", "ペイジェント", "GMO-PG", "SMBCファイナンス", "ヤマトフィナンシャル", "ゼウス", "SMBC-GMO", "GMO-MG", "三菱UFJニコス", "NTTコムオンライン", "楽天カード", "pay.jp"]

psp_dic = {
    "DGFT" : ["オーソリ", "オーソリ売上", "カードチェック", "カード預かり登録", "カード預かり更新"], 
    "ソニーペイメント" : ["オーソリ", "オーソリ売上", "カードチェック", "カード預かり登録", "カード預かり更新"],
    "ペイジェント" : ["オーソリ", "オーソリ売上", "カードチェック", "カード預かり登録", "カード預かり更新"], 
    "GMO-PG" : ["オーソリ", "オーソリ売上", "カードチェック", "カード預かり登録", "カード預かり更新"],
    "SMBCファイナンス" : ["オーソリ", "オーソリ売上", "カードチェック", "カード預かり(都度課金)登録", "カード預かり(都度課金)更新", "カード預かり(継続課金)登録", "カード預かり(継続課金)更新"], 
    "ヤマトフィナンシャル" : ["オーソリ", "オーソリ売上", "カードチェック", "カード預かり登録", "カード預かり更新"],
    "ゼウス" : ["オーソリ", "オーソリ売上", "カードチェック", "カード預かり登録", "カード預かり更新"],
    "SMBC-GMO" : ["オーソリ", "オーソリ売上", "カードチェック", "カード預かり登録", "カード預かり更新"],
    "GMO-MG" : ["カード預かり登録"],
    "三菱UFJニコス" : ["オーソリ", "オーソリ売上", "カードチェック"],
    "NTTコムオンライン" : ["カードチェック"],
    "楽天カード" : ["オーソリ", "オーソリ売上", "カードチェック", "カード預かり登録"],
    "pay.jp" : ["オーソリ", "オーソリ売上", "カード預かり登録"]
}

options = ["オプション全表示", "オプション全非表示"]

input_methods = ["手動入力", "磁気", "IC"]

modes = ["スタンドアロン", "PC 連動"]

content = "全パラメータをバリデーションルールに従って入力し、決済が正常に終了すること"

message = "取引履歴に記録されていることを確認する"

def do_loop(ws):
    num = 14

    for mode in modes:
        for psp in psp_list:
            methods = psp_dic[psp]
            for method in methods:
                for option in options:
                    for input_method in input_methods:
                        title = mode + "\n　" + psp + "\n　　" + method + "\n　　　" + option + "\n　　　　" + input_method
                        ws["C" + str(num)] = title
                        ws["I" + str(num)] = "正常系"
                        ws["K" + str(num)] = content
                        ws["S" + str(num)] = message
                        num += 1

def report():
    wb = openpyxl.load_workbook("test.xlsx")
    ws = wb['決済']  # ワークシートを指定
    ws = wb.active  # アクティブなワークシートを選択
    
    do_loop(ws)

    #ws["C14"] = "test"
    #ws["I14"] = "正常系"
    #ws["Q14"] = "test1"
    #ws["AA14"] = "test2"

    wb.save('test.xlsx')

if __name__ == '__main__':
    report()
