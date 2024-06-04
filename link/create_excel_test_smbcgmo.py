# -*- coding: utf-8 -*-

# エビデンスの画像をエクセル形式でまとめる

import openpyxl
from openpyxl import Workbook

device_list = ["Pay TG PC アプリケーション", "Smart TG PC アプリケーション", "Smart TG スタンドアロン"]

check_list = [("有効性チェックなし", "有効性チェックが呼ばれないこと"), ("有効性チェックなし", "有効性チェックが呼ばれること")]

input_methods = [
     ("", "存在しない会員IDを指定し、カード預かり登録を実行する" , "カードが新規に登録されること"),
     ("", "存在する会員IDを指定し、カード預かり登録を実行する", "カードが新規に登録されること"),
     ("", "クレジットカードを登録していないが、存在する会員IDを指定し、カード登録連番0を指定して、カード預かり登録を実行する", "カードが新規に登録されること"),
     ("", "クレジットカードを登録してる既存の会員IDを指定し、カード登録連番0を指定して、カード預かり登録を実行する", "カードが更新されること"),
]


content = "全パラメータをバリデーションルールに従って入力し、決済が正常に終了すること"

message = "取引履歴に記録されていることを確認する"

def do_loop(ws):
    num = 14

    for device in device_list:
        for check in check_list:
            for input_method in input_methods:
                title = "SMBC-GMO カード預かり登録\n" + device + "\n" + "　" + check[0] + "\n　" + input_method[0]
                ws["C" + str(num)] = title
                ws["I" + str(num)] = "正常系"
                ws["S" + str(num)] = input_method[1]
                ws["AI" + str(num)] = input_method[2] + "\n" + check[1]
                num += 1

def report():
    wb = openpyxl.load_workbook("test.xlsx")
    ws = wb['決済']  # ワークシートを指定
    ws = wb.active  # アクティブなワークシートを選択
    
    do_loop(ws)

    #ws["C14"] = "test"
    #ws["I14"] = "正常系"
    #ws["Q14"] = "test1"
    #ws["AA14"] = "test2"

    wb.save('test.xlsx')

if __name__ == '__main__':
    report()
