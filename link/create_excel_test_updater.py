# -*- coding: utf-8 -*-

# エビデンスの画像をエクセル形式でまとめる

import openpyxl
from openpyxl import Workbook

windows = ["Windows11 Pro", "Windows10 Home"]

devices = ["VEGA3000", "P2 Pro", "P2 Lite SE"]

devices2 = ["VEGA3000", "P2 Lite SE"]

versions = ["1.0", ""]

folders = ["デスクトップにアップデータを配置してで実行", "Users フォルダ配下以外にアップデータを配置してで実行"]

normal = [
	[
	  "アップデータ\n　{device}\n　{windows}\n　ダウンロードフォルダにアップデータを配置してで実行",
	  "正常系",
	  "Sqlite3コマンドが使用できないPCで実施する\nPC アプリのバージョン（V{version}）をインストーラでインストールする",
	  "アプリを一度も起動せず（SQLite3 DB が作成されていない状態で）、アップデータから最新版にアップデートする\nケースで指定した操作を実行する。",
	  "アップデート後起動できること\nアップデート後、自動ペアリングされないこと\nバージョンが最新になっていること\n取引を実行し、正常に記録されること"
	],
	[
	  "アップデータ\n　{device}\n　{windows}",
	  "正常系",
	  "Sqlite3コマンドが使用できないPCで実施する\nPC アプリのバージョン（V{version}）をインストーラでインストールする",
	  "アプリを一度起動した状態（SQLite3 DB が作成されているが取引履歴がない状態で）、アップデータから最新版にアップデートする\nケースで指定した操作を実行する。",
	  "アップデート後起動できること\nアップデート後自動ペアリングできること\nバージョンが最新になっていること\n取引を実行し、正常に記録されること\nSQLite コマンドで SQLite DB の中を確認し、他のテーブル含めデータが正しく引き継げていること"
	],
	[
	  "アップデータ\n　{device}\n　{windows}",
	  "正常系",
	  "Sqlite3コマンドが使用できないPCで実施する\nPC アプリのバージョン（V{version}）をインストーラでインストールする\n取引を実行し、取引履歴があること",
	  "アップデータから最新版にアップデートする\nケースで指定した操作を実行する。",
	  "アップデート後起動できること\nアップデート後自動ペアリングできること\nバージョンが最新になっていること\n取引を実行し、正常に記録されること\nSQLite コマンドで SQLite DB の中を確認し、他のテーブル含めデータが正しく引き継げていること"
	],
]

abnormal = [
	[
	  "アップデータ\n　{device}\n　{windows}",
	  "異常系",
	  "Sqlite3コマンドが使用できないPCで実施する\nアップデータを展開したフォルダから「sqlite3.exe」を削除する",
	  "アップデータを実行する",
	  "アップデートに失敗すること\nアップデート失敗後も起動できること\nアップデート後自動ペアリングできないこと\nアップデート前のバージョンになっていること\n取引を実行し、正常に記録されること"
	],
	[
	  "アップデータ\n　{device}\n　{windows}",
	  "異常系",
	  "Sqlite3コマンドが使用できないPCで実施する\nアップデータを展開したフォルダから「sqlite3.exe」を削除する",
	  "アプリを一度起動した状態（SQLite3 DB が作成されているが取引履歴がない状態で）、アップデータから最新版にアップデートする\n　ケースで指定した操作を実行する。",
	  "アップデートに失敗すること\nアップデート失敗後も起動できること\nアップデート後自動ペアリングできること\nアップデート前のバージョンになっていること\n取引を実行し、正常に記録されること"
	],
	[
	  "アップデータ\n　{device}\n　{windows}",
	  "異常系",
	  "Sqlite3コマンドが使用できないPCで実施する\n取引を実行し、取引履歴があること\nアップデータを展開したフォルダから「sqlite3.exe」を削除する",
	  "アップデータから最新版にアップデートする\nケースで指定した操作を実行する。",
	  "アップデートに失敗すること\nアップデート失敗後も起動できること\nアップデート後自動ペアリングできること\nアップデート前のバージョンになっていること\n取引を実行し、正常に記録されること"
	],
]

folder_pattern = [
	[
	  "アップデータ\n　{device}\n　{windows}\n　{folder}",
	  "正常系",
	  "Sqlite3コマンドが使用できないPCで実施する\nPC アプリのバージョン（V{version}）をインストーラでインストールする\n取引を実行し、取引履歴があること",
	  "アップデータから最新版にアップデートする\nケースで指定した操作を実行する。",
	  "アップデート後起動できること\nアップデート後自動ペアリングできること\nバージョンが最新になっていること\n取引を実行し、正常に記録されること\nSQLite コマンドで SQLite DB の中を確認し、他のテーブル含めデータが正しく引き継げていること"
	],
]

def do_loop(ws):
    num = 14

    for w in windows:
        msg = "msgコマンドが使用できることを確認する\n"
        if w == "Windows10 Home":
            msg = "msgコマンドが使用できないことを確認する\n"
        for pattern in normal:
            for d in devices:
                for v in versions:
                    ws["C" + str(num)] = pattern[0].format(device=d, windows=w)   # テスト項目（評価内容）
                    ws["I" + str(num)] = pattern[1]   # 分類
                    ws["K" + str(num)] = "機能テスト"   # テスト観点種別
                    if len(v) == 0:
                        if d == "VEGA3000":
                            v = "3.15"
                        else:
                            v = "1.11"
                    ws["N" + str(num)] = msg + pattern[2].format(version=v)   # 事前条件
                    ws["V" + str(num)] = pattern[3]   # 操作及び入力値
                    ws["AL" + str(num)] = pattern[4]   # 振る舞い
                    ws["BI" + str(num)] = "1" # テスト実施
                    num += 1
        
        for pattern in abnormal:
            for d in devices:
                ws["C" + str(num)] = pattern[0].format(device=d, windows=w)   # テスト項目（評価内容）
                ws["I" + str(num)] = pattern[1]   # 分類
                ws["K" + str(num)] = "機能テスト"   # テスト観点種別
                ws["N" + str(num)] = pattern[2]    # 事前条件
                ws["V" + str(num)] = pattern[3]   # 操作及び入力値
                ws["AL" + str(num)] = pattern[4]   # 振る舞い
                ws["BI" + str(num)] = "1" # テスト実施
                num += 1

    for w in windows:
        msg = "msgコマンドが使用できることを確認する\n"
        if w == "Windows10 Home":
            msg = "msgコマンドが使用できないことを確認する\n"
        for pattern in folder_pattern:
            for f in folders:
                for d in devices2:
                    for v in versions:
                        ws["C" + str(num)] = pattern[0].format(device=d, windows=w, folder=f)   # テスト項目（評価内容）
                        ws["I" + str(num)] = pattern[1]   # 分類
                        ws["K" + str(num)] = "機能テスト"   # テスト観点種別
                        if len(v) == 0:
                            if d == "VEGA3000":
                                v = "3.15"
                            else:
                                v = "1.11"
                        ws["N" + str(num)] = msg + pattern[2].format(version=v)   # 事前条件
                        ws["V" + str(num)] = pattern[3]   # 操作及び入力値
                        ws["AL" + str(num)] = pattern[4]   # 振る舞い
                        ws["BI" + str(num)] = "1" # テスト実施
                        num += 1

def report():
    wb = openpyxl.load_workbook("test.xlsx")
    ws = wb['決済']  # ワークシートを指定
    ws = wb.active  # アクティブなワークシートを選択
    
    do_loop(ws)

    wb.save('test.xlsx')

if __name__ == '__main__':
    report()
