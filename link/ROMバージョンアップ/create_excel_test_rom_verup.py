# -*- coding: utf-8 -*-

# エビデンスの画像をエクセル形式でまとめる

import openpyxl
from openpyxl import Workbook

device = ["P2 Pro", "P2 Lite SE"]

mode = ["PC 連動", "スタンドアロン"]

rom = ["ROM バージョンアップ前", "ROM バージョンアップ後"]

communication = ["SIM（docomo）\n　PC 連動", "SIM（softbank）\n　スタンドアロン", "Wi-Fi\n　PC 連動"]

test_pattern = [
	[
	  "{0}\n　{1}\n　{2}\n　　システム更新",
	  "正常系",
	  "UIテスト",
	  "",
	  "端末を起動し、ホーム画面からメニューを開く",
	  "メニューに「システム更新」が表示されること"
	],
	[
	  "{0}\n　{1}\n　{2}\n　　システム更新",
	  "正常系",
	  "UIテスト",
	  "",
	  "ホーム画面のメニューから「システム更新」を押下する",
	  "「はい」「いいえ」のダイアログが表示されること"
	],
	[
	  "{0}\n　{1}\n　{2}\n　　システム更新",
	  "正常系",
	  "UIテスト",
	  "",
	  "ホーム画面のメニューから「システム更新」を押下した後のダイアログで「はい」を押下する",
	  "「システム更新には、大量のデータ通信が発生します。Wi-Fi環境での操作を推奨します。」ダイアログが表示されること"
	],
	[
	  "{0}\n　{1}\n　{2}\n　　システム更新",
	  "正常系",
	  "UIテスト",
	  "ホーム画面のメニューから「システム更新」を押下した後のダイアログで「はい」を押下する",
	  "システム更新には、大量のデータ通信が発生します。Wi-Fi環境での操作を推奨します。で「はい」を押下する",
	  "ROMのバージョンアップ画面に切り替わること"
	],
	[
	  "{0}\n　{1}\n　{2}\n　　システム更新",
	  "正常系",
	  "UIテスト",
	  "ホーム画面のメニューから「システム更新」を押下した後のダイアログで「はい」を押下する",
	  "システム更新には、大量のデータ通信が発生します。Wi-Fi環境での操作を推奨します。で「いいえ」を押下する",
	  "ROMのバージョンアップ画面に切り替わらず、ホーム画面に戻ること"
	],
	[
	  "{0}\n　{1}\n　{2}\n　　システム更新",
	  "正常系",
	  "UIテスト",
	  "ホーム画面のメニューから「システム更新」を押下した後のダイアログで「いいえ」を押下する",
	  "ホーム画面のメニューから「システム更新」を押下した後のダイアログで「いいえ」を押下する",
	  "ROMのバージョンアップ画面に切り替わらず、ホーム画面に戻ること"
	],
]

test_pattern2 = [
	[
	  "{0}\n　{1}\n　　システム更新",
	  "正常系",
	  "UIテスト",
	  "",
	  "ROM のバージョンアップ画面でバージョンアップを実施する",
	  "正常にバージョンアップできること"
	],
]

test_pattern3 = [
	[
	  "{0}\n　{1}\n　{2}\n　　Wi-Fi 設定",
	  "正常系",
	  "UIテスト",
	  "TMSWEB 管理画面の端末設定画面で、LTE 自動切り替え(Pay TG) / Wi-Fi 許可(Smart TG)をLTE 許可(Pay TG) / Wi-Fi 許可しない(Smart TG)に設定する",
	  "決済端末を再起動し、ホーム画面からメニューを開く",
	  "Wi-Fi 設定が表示されないこと"
	],
	[
	  "{0}\n　{1}\n　{2}\n　　Wi-Fi 設定",
	  "正常系",
	  "UIテスト",
	  "TMSWEB 管理画面の端末設定画面で、LTE 自動切り替え(Pay TG) / Wi-Fi 許可(Smart TG)をLTE/3G 自動切替(Pay TG) / Wi-Fi 許可する(Smart TG)に設定する",
	  "決済端末を再起動し、ホーム画面からメニューを開く",
	  "Wi-Fi 設定が表示されること"
	],
	[
	  "{0}\n　{1}\n　{2}\n　　Wi-Fi 設定",
	  "正常系",
	  "UIテスト",
	  "TMSWEB 管理画面の端末設定画面で、LTE 自動切り替え(Pay TG) / Wi-Fi 許可(Smart TG)をLTE/3G 自動切替(Pay TG) / Wi-Fi 許可する(Smart TG)に設定する",
	  "ホーム画面のメニューから「Wi-Fi設定」を押下する",
	  "「はい」「いいえ」のダイアログが表示されること"
	],
	[
	  "{0}\n　{1}\n　{2}\n　　Wi-Fi 設定",
	  "正常系",
	  "UIテスト",
	  "TMSWEB 管理画面の端末設定画面で、LTE 自動切り替え(Pay TG) / Wi-Fi 許可(Smart TG)をLTE/3G 自動切替(Pay TG) / Wi-Fi 許可する(Smart TG)に設定する",
	  "ホーム画面のメニューから「Wi-Fi設定」を押下した後のダイアログで「はい」を押下する",
	  "Wi-Fi 設定画面に切り替わること"
	],
	[
	  "{0}\n　{1}\n　{2}\n　　Wi-Fi 設定",
	  "正常系",
	  "UIテスト",
	  "TMSWEB 管理画面の端末設定画面で、LTE 自動切り替え(Pay TG) / Wi-Fi 許可(Smart TG)をLTE/3G 自動切替(Pay TG) / Wi-Fi 許可する(Smart TG)に設定する",
	  "ホーム画面のメニューから「Wi-Fi設定」を押下した後のダイアログで「いいえ」を押下する",
	  "Wi-Fi 設定画面に切り替わらず、ホーム画面が表示されること"
	],
	[
	  "{0}\n　{1}\n　{2}\n　　Wi-Fi 設定",
	  "正常系",
	  "UIテスト",
	  "TMSWEB 管理画面の端末設定画面で、LTE 自動切り替え(Pay TG) / Wi-Fi 許可(Smart TG)をLTE/3G 自動切替(Pay TG) / Wi-Fi 許可する(Smart TG)に設定する",
	  "ホーム画面のメニューから「Wi-Fi設定」を押下した後のダイアログで「Wi-Fi 設定」を押下し、Wi-Fi 設定をした後、再起動する",
	  "Wi-Fi が OFF になっていること"
	],
	[
	  "{0}\n　{1}\n　{2}\n　　Wi-Fi 設定",
	  "正常系",
	  "UIテスト",
	  "TMSWEB 管理画面の端末設定画面で、LTE 自動切り替え(Pay TG) / Wi-Fi 許可(Smart TG)をLTE/3G 自動切替(Pay TG) / Wi-Fi 許可する(Smart TG)に設定する",
	  "ホーム画面のメニューから「Wi-Fi設定」を押下した後のダイアログで「Wi-Fi 設定」を押下し、Wi-Fi 設定をした後、再起動し、さらにもう一度再起動する",
	  "Wi-Fi が OFF になっていること"
	],
	[
	  "{0}\n　{1}\n　{2}\n　　Wi-Fi 設定",
	  "正常系",
	  "UIテスト",
	  "TMSWEB 管理画面の端末設定画面で、LTE 自動切り替え(Pay TG) / Wi-Fi 許可(Smart TG)をLTE/3G 自動切替(Pay TG) / Wi-Fi 許可する(Smart TG)に設定する",
	  "ホーム画面のメニューから「Wi-Fi設定」を押下した後のダイアログで「Wi-Fi 設定」を押下し、Wi-Fi 設定をした後、決済を実行する",
	  "決済が実行できないこと\n　エラーダイアログが表示されること"
	],
	[
	  "{0}\n　{1}\n　{2}\n　　Wi-Fi 設定",
	  "正常系",
	  "UIテスト",
	  "TMSWEB 管理画面の端末設定画面で、LTE 自動切り替え(Pay TG) / Wi-Fi 許可(Smart TG)をLTE/3G 自動切替(Pay TG) / Wi-Fi 許可する(Smart TG)に設定する",
	  "ホーム画面のメニューから「Wi-Fi設定」を押下した後のダイアログで「Wi-Fi 設定」を押下し、Wi-Fi 設定をした後、再度オフに設定して、決済を実行する",
	  "決済が実行できること"
	],
]

def do_loop(ws):
    num = 14

    for pattern in test_pattern:
        for d in device:
            for m in mode:
                for r in rom:
                    ws["C" + str(num)] = pattern[0].format(d, m, r)   # テスト項目（評価内容）
                    ws["I" + str(num)] = pattern[1]   # 分類
                    ws["K" + str(num)] = pattern[2]   # テスト観点種別
                    ws["N" + str(num)] = pattern[3]   # 事前条件
                    ws["V" + str(num)] = pattern[4]   # 操作および入力値
                    ws["AL" + str(num)] = pattern[5]   # 振る舞い
                    ws["BI" + str(num)] = "1"
                    num += 1

    for pattern in test_pattern2:
        for d in device:
            for m in communication:
                ws["C" + str(num)] = pattern[0].format(d, m)   # テスト項目（評価内容）
                ws["I" + str(num)] = pattern[1]   # 分類
                ws["K" + str(num)] = pattern[2]   # テスト観点種別
                ws["N" + str(num)] = pattern[3]   # 事前条件
                ws["V" + str(num)] = pattern[4]   # 操作および入力値
                ws["AL" + str(num)] = pattern[5]   # 振る舞い
                ws["BI" + str(num)] = "1"
                num += 1

    for pattern in test_pattern3:
        for d in device:
            for m in mode:
                for r in rom:
                    ws["C" + str(num)] = pattern[0].format(d, m, r)   # テスト項目（評価内容）
                    ws["I" + str(num)] = pattern[1]   # 分類
                    ws["K" + str(num)] = pattern[2]   # テスト観点種別
                    ws["N" + str(num)] = pattern[3]   # 事前条件
                    ws["V" + str(num)] = pattern[4]   # 操作および入力値
                    ws["AL" + str(num)] = pattern[5]   # 振る舞い
                    ws["BI" + str(num)] = "1"
                    num += 1

def report():
    wb = openpyxl.load_workbook("test.xlsx")
    ws = wb['ROMのバージョンアップ対応']  # ワークシートを指定
    ws = wb.active  # アクティブなワークシートを選択
    
    do_loop(ws)

    wb.save('test.xlsx')

if __name__ == '__main__':
    report()
