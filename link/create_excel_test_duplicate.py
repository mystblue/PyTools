# -*- coding: utf-8 -*-

# エビデンスの画像をエクセル形式でまとめる

import openpyxl
from openpyxl import Workbook

device_list = ["P2Pro", "VEGA3000"]

check_list = ["英字を含む15桁のシリアル番号", "英字を含む15桁未満のシリアル番号"]

input_methods = [
     ("",
      "%sを新規に登録する",
      "正常に登録できること"
     ),
     ("%sを登録しておく",
      "事前条件と同じシリアル番号を新規に登録する",
      "バリデーションエラーとなること"
     ),
     ("%sを登録しておく",
      "端末設定の編集画面を開き、内容を修正して、登録、確認、更新する",
      "正常に更新できること"
     ),
     ("%sを端末設定画面から削除する",
      "削除したシリアル番号と同じシリアル番号を登録する",
      "正常に登録できること"
     ),
     ("",
      "削除して再登録したシリアル番号と同じシリアル番号を登録する",
      "バリデーションエラーとなること"
     ),
]


def do_loop(ws):
    num = 14

    for device in device_list:
        for check in check_list:
            for input_method in input_methods:
                title = device
                ws["C" + str(num)] = title
                ws["I" + str(num)] = "正常系"
                ws["K" + str(num)] = input_method[0].replace("%s", check)
                ws["S" + str(num)] = input_method[1].replace("%s", check)
                ws["AI" + str(num)] = input_method[2]
                num += 1

def report():
    wb = openpyxl.load_workbook("test.xlsx")
    ws = wb['決済']  # ワークシートを指定
    ws = wb.active  # アクティブなワークシートを選択
    
    do_loop(ws)

    #ws["C14"] = "test"
    #ws["I14"] = "正常系"
    #ws["Q14"] = "test1"
    #ws["AA14"] = "test2"

    wb.save('test.xlsx')

if __name__ == '__main__':
    report()
