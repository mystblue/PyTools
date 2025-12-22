# -*- coding: utf-8 -*-

# エビデンスの画像をエクセル形式でまとめる

import openpyxl
from openpyxl import Workbook

device = ["P2 Pro", "P2 Lite SE"]

modes = ["自動ペアリング", "手動ペアリング"]

pc_rendou = [
	[
	  "\n　",
	  "システムテスト",
	  "1.0.53 にアップデートする",
	  "アップデート後のペアリング状態で一度USBケーブルを抜き、再度接続する",
	  "{0}できること"
	],
	[
	  "\n　",
	  "システムテスト",
	  "1.0.53 にアップデートする",
	  "USB接続状態でPCアプリをタスクトレイから一度終了させ、再起動する",
	  "{0}できること"
	],
	[
	  "\n　",
	  "システムテスト",
	  "1.0.53 にアップデートする",
	  "USB未接続状態でPCアプリをタスクトレイから一度終了させ、再起動する",
	  "{0}できること"
	],
	[
	  "\n　",
	  "システムテスト",
	  "1.0.53 にアップデートする",
	  "USB接続状態でPCをシャットダウンし、再起動する",
	  "PC アプリが自動起動し、その後{0}できること"
	],
	[
	  "\n　",
	  "システムテスト",
	  "1.0.53 にアップデートする",
	  "USB未接続状態でPCをシャットダウンし、再起動する",
	  "PC アプリが自動起動し、その後{0}できること"
	],
	[
	  "\n　",
	  "システムテスト",
	  "1.0.53 にアップデートする",
	  "USB接続状態で端末を再起動させる",
	  "{0}できること"
	],
	[
	  "\n　",
	  "システムテスト",
	  "1.0.53 にアップデートする",
	  "USB未接続状態で端末を再起動させる",
	  "{0}できること"
	],
]

def do_loop(ws):
    num = 14

    for d in device:
        for mode in modes:
            for pattern in pc_rendou:
                ws["C" + str(num)] = d + pattern[0] + mode   # テスト項目（評価内容）
                ws["I" + str(num)] = "正常系"   # 分類
                ws["K" + str(num)] = pattern[1]   # テスト観点種別
                ws["N" + str(num)] = pattern[2]   # 事前条件
                ws["V" + str(num)] = pattern[3]   # 操作および入力値
                ws["AL" + str(num)] = pattern[4].format(mode)   # 振る舞い
                ws["BI" + str(num)] = "1"
                num += 1

def report():
    wb = openpyxl.load_workbook("test.xlsx")
    ws = wb['自動インストール']  # ワークシートを指定
    ws = wb.active  # アクティブなワークシートを選択
    
    do_loop(ws)

    wb.save('test.xlsx')

if __name__ == '__main__':
    report()
