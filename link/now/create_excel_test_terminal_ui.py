# -*- coding: utf-8 -*-

# エビデンスの画像をエクセル形式でまとめる

import openpyxl
from openpyxl import Workbook

device = ["P2Pro（Windows10 32bit）", "P2Pro（Windows10 64bit）", "P2Pro（Windows11 64bit）", "P2 Lite SE（Windows10 32bit）", "P2 Lite SE（Windows10 64bit）", "P2 Lite SE（Windows11 64bit）"]

pc_rendou = [
    # 端末接続時の初期認証
	[
	  "\n　初回認証",
	  "機能テスト\nログ・モニタリングテスト",
	  "PC 連動モードで PC アプリを新規にインストールする",
	  "PC と決済端末を USB ケーブルで接続し、PC アプリからポートを選択して、接続ボタンをクリックする。",
	  "端末でペアリングパスワードを正しく入力し、ペアリングに成功すること\nログに「端末初期認証結果：True」「ペアリング結果：True」と出力されること"
	],
	[
	  "\n　別ポートでの接続",
	  "機能テスト\nログ・モニタリングテスト",
	  "初回認証済み",
	  "PC と決済端末を初回認証したのと別の USB ポートへ接続し、PC アプリからポートを選択して、接続ボタンをクリックする。",
	  "端末でペアリングパスワードを正しく入力し、ペアリングに成功すること\nログに「ペアリング結果：True」と出力されること"
	],
	[
	  "\n　異なる端末",
	  "機能テスト\nログ・モニタリングテスト",
	  "初回認証済み",
	  "PC と決済端末を初回認証したのと別の決済端末を接続する",
	  "自動ペアリングに失敗すること\nログに「自動接続を試みます。」が出力され、「ペアリング結果：True」と出力されないこと"
	],
	[
	  "\n　異なる端末",
	  "機能テスト\nログ・モニタリングテスト",
	  "初回認証済み",
	  "PC と決済端末を初回認証したのと別の決済端末を接続し、PC アプリからポートを選択して、接続ボタンをクリックする。",
	  "端末でペアリングパスワードを正しく入力し、ペアリングに成功すること\nログに「ペアリング結果：True」と出力されること"
	],
	# クライアント証明書変更
	[
	  "\n　クライアント証明書変更",
	  "システムテスト",
	  "初回認証済み",
	  "TMSWEB 管理画面でクライアント証明書を変更し、決済端末を再起動する。\nPC と決済端末を接続し、PC アプリからポートを選択して、接続ボタンをクリックする。",
	  "端末でペアリングパスワードを正しく入力し、ペアリングに成功すること\nログに「ペアリング結果：True」と出力されること"
	],
	# 決済
	[
	  "\n　決済",
	  "システムテスト",
	  "初回認証後",
	  "決済を行う。",
	  "決済に成功すること\n取引結果が表示されること\nTMSWEB 管理画面に取引履歴が記録されること\nログにリクエストとレスポンスが出力されること"
	],
	# 間違ったペアリングパスワード
	[
	  "\n　間違ったペアリングパスワード",
	  "機能テスト\nログ・モニタリングテスト",
	  "PC 連動モードで PC アプリを新規にインストールする",
	  "PC と決済端末を USB ケーブルで接続し、PC アプリからポートを選択して、接続ボタンをクリックする。",
	  "端末で謝ったペアリングパスワードを入力し、ペアリングに失敗すること"
	],
	# 手動ペアリング失敗
	[
	  "\n　手動ペアリング中の切断",
	  "P2 Pro / P2 Lite SE",
	  "",
	  "・PCアプリケーションで決済端末が接続されているCOMポートを指定し端末接続を実施する。\n・ペアリングパスワード入力画面で決済端末とPCの接続を切断する。",
	  "端末手動接続ペアリングに失敗すること。"
	],
	# 自動接続
	[
	  "\n　自動接続",
	  "機能テスト\nログ・モニタリングテスト",
	  "初回認証済み",
	  "PC と決済端末がペアリングされている状態で、USB ケーブルを抜いて、再度接続する。",
	  "自動ペアリングに成功すること\nログに「COM○○ポートに接続されている決済端末が操作可能となりました。」と出力されること"
	],
	[
	  "\n　自動接続（PC を先に起動）",
	  "機能テスト\nログ・モニタリングテスト",
	  "",
	  "PCと決済端末の電源をOFFにした状態でPCを先に起動し、その後決済端末を起動する",
	  "自動ペアリングに成功すること\nログに「COM○○ポートに接続されている決済端末が操作可能となりました。」と出力されること"
	],
	[
	  "\n　自動接続（決済端末を先に起動）",
	  "機能テスト\nログ・モニタリングテスト",
	  "",
	  "PCと決済端末の電源をOFFにした状態で決済端末を先に起動し、その後PCを起動する",
	  "自動ペアリングに成功すること\nログに「COM○○ポートに接続されている決済端末が操作可能となりました。」と出力されること"
	],
	# 自動接続後の決済
	[
	  "\n　自動接続後の決済",
	  "システムテスト",
	  "",
	  "PCと決済端末を USB ケーブルを抜き差しして自動ペアリングされた後に決済を行う",
	  "決済に成功すること\n取引結果が表示されること\nTMSWEB 管理画面に取引履歴が記録されること\nログにリクエストとレスポンスが出力されること"
	],
	[
	  "\n　自動接続（異なるポートで実施）",
	  "機能テスト\nログ・モニタリングテスト",
	  "初回認証済み",
	  "PC と決済端末がペアリングされている状態で、USB ケーブルを抜いて、異なる USB ポートに接続する。",
	  "自動ペアリングが実行されないこと\nログに「自動接続を試みます。」と出力されないこと"
	],
	# 自動接続中の切断
	[
	  "\n　自動接続中の切断",
	  "ログ・モニタリングテスト\nエラーハンドリングテスト",
	  "",
	  "PCと決済端末を USB ケーブルを抜き差しして自動ペアリング中に USB ケーブルを抜く",
	  "自動ペアリングに失敗すること"
	],
	# P2Pro と P2 Lite SE を同時ペアリング
	[
	  "\n　P2Pro をペアリングした後に P2Pro をペアリング",
	  "機能テスト\nログ・モニタリングテスト",
	  "P2Pro をペアリングする",
	  "P2Pro をペアリングした後に、P2Pro をペアリングする",
	  "ペアリングできないこと"
	],
	[
	  "\n　P2Pro をペアリングした後に P2 Lite SE をペアリング",
	  "機能テスト\nログ・モニタリングテスト",
	  "P2Pro をペアリングする",
	  "P2Pro をペアリングした後に、P2 Lite SE をペアリングする",
	  "ペアリングできないこと"
	],
	[
	  "\n　P2 Lite SE をペアリングした後に P2Pro をペアリング",
	  "機能テスト\nログ・モニタリングテスト",
	  "P2 Lite SE をペアリングする",
	  "P2 Lite SE をペアリングした後に、P2Pro をペアリングする",
	  "ペアリングできないこと"
	],
	[
	  "\n　P2 Lite SE をペアリングした後に P2 Lite SE をペアリング",
	  "機能テスト\nログ・モニタリングテスト",
	  "P2 Lite SE をペアリングする",
	  "P2 Lite SE をペアリングした後に、P2 Lite SE をペアリングする",
	  "ペアリングできないこと"
	],
]

def do_loop(ws):
    num = 14

    for pattern in pc_rendou:
        for d in device:
            ws["C" + str(num)] = d + pattern[0]   # テスト項目（評価内容）
            ws["I" + str(num)] = "正常系"   # 分類
            ws["K" + str(num)] = pattern[1]   # テスト観点種別
            ws["N" + str(num)] = pattern[2]   # 事前条件
            ws["V" + str(num)] = pattern[3]   # 操作および入力値
            ws["AL" + str(num)] = pattern[4]   # 振る舞い
            ws["BI" + str(num)] = "1"
            num += 1

def report():
    wb = openpyxl.load_workbook("test.xlsx")
    ws = wb['決済端末のペアリング']  # ワークシートを指定
    ws = wb.active  # アクティブなワークシートを選択
    
    do_loop(ws)

    wb.save('test.xlsx')

if __name__ == '__main__':
    report()