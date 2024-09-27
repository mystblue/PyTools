# -*- coding: utf-8 -*-

# エビデンスの画像をエクセル形式でまとめる

import openpyxl
from openpyxl import Workbook

device = ["P2 Lite SE PC 連動", "P2 Lite SE スタンドアロン"]

pc_rendou = [
	[
	  "\n　キーボードの変更",
	  "日本語のキーボードが使えるように設定を変更し、デフォルトを英語キーボードにしておく",
	  "キーボードを日本語に変更する",
	  "日本語キーボードに変更できること"
	],
	[
	  "\n　音量の変更",
	  "",
	  "音量を変更し任意の取引種別を実施する。音量は最小にして、段階的に最大まであげて確認する",
	  "音量に合わせて操作音が変化すること"
	],
	[
	  "\n　時刻同期",
	  "",
	  "決済端末を起動し時刻を確認する。",
	  "正常な時刻が表示されることを確認する。"
	],
	[
	  "\n　災害警報",
	  "",
	  "決済端末を起動状態で災害警報を待つ。",
	  "災害警報が表示されること。\nOKボタンを押下してアプリの画面に戻ること。"
	],
]

def do_loop(ws):
    num = 14

    for pattern in pc_rendou:
        for d in device:
            ws["C" + str(num)] = d + pattern[0]   # テスト項目（評価内容）
            ws["I" + str(num)] = "正常系"   # 分類
            ws["K" + str(num)] = pattern[1]   # テスト観点種別
            ws["N" + str(num)] = pattern[2]   # 事前条件
            ws["V" + str(num)] = pattern[3]   # 操作および入力値
            ws["AL" + str(num)] = pattern[4]   # 振る舞い
            ws["BI" + str(num)] = "1"
            num += 1

def report():
    wb = openpyxl.load_workbook("test.xlsx")
    ws = wb['レシート']  # ワークシートを指定
    ws = wb.active  # アクティブなワークシートを選択
    
    do_loop(ws)

    wb.save('test.xlsx')

if __name__ == '__main__':
    report()
