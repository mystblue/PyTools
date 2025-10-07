# -*- coding: utf-8 -*-

# エビデンスの画像をエクセル形式でまとめる

import openpyxl
from openpyxl import Workbook

def do_loop(ws):

    #num = 5
    for num in range(4, 14):
        ws["D" + str(num)] = "=COUNTIFS(リンク様_SIM不良対象リスト!K2:K331, A" + str(num) + ",リンク様_SIM不良対象リスト!M2:M331, D2)"
        ws["E" + str(num)] = "=COUNTIFS(リンク様_SIM不良対象リスト!K2:K331, A" + str(num) + ",リンク様_SIM不良対象リスト!M2:M331, E2)"
        ws["F" + str(num)] = "=COUNTIFS(リンク様_SIM不良対象リスト!K2:K331, A" + str(num) + ",リンク様_SIM不良対象リスト!M2:M331, F2)"
        ws["G" + str(num)] = "=COUNTIFS(リンク様_SIM不良対象リスト!K2:K331, A" + str(num) + ",リンク様_SIM不良対象リスト!M2:M331, G2)"
        ws["H" + str(num)] = "=COUNTIFS(リンク様_SIM不良対象リスト!K2:K331, A" + str(num) + ",リンク様_SIM不良対象リスト!M2:M331, H2)"
        ws["I" + str(num)] = "=COUNTIFS(リンク様_SIM不良対象リスト!K2:K331, A" + str(num) + ",リンク様_SIM不良対象リスト!M2:M331, I2)"
        ws["J" + str(num)] = "=COUNTIFS(リンク様_SIM不良対象リスト!K2:K331, A" + str(num) + ",リンク様_SIM不良対象リスト!M2:M331, J2)"
        ws["K" + str(num)] = "=COUNTIFS(リンク様_SIM不良対象リスト!K2:K331, A" + str(num) + ",リンク様_SIM不良対象リスト!M2:M331, K2)"
        ws["L" + str(num)] = "=COUNTIFS(リンク様_SIM不良対象リスト!K2:K331, A" + str(num) + ",リンク様_SIM不良対象リスト!M2:M331, L2)"
        ws["M" + str(num)] = "=COUNTIFS(リンク様_SIM不良対象リスト!K2:K331, A" + str(num) + ",リンク様_SIM不良対象リスト!M2:M331, M2)"
        ws["N" + str(num)] = "=COUNTIFS(リンク様_SIM不良対象リスト!K2:K331, A" + str(num) + ",リンク様_SIM不良対象リスト!M2:M331, N2)"
        ws["O" + str(num)] = "=COUNTIFS(リンク様_SIM不良対象リスト!K2:K331, A" + str(num) + ",リンク様_SIM不良対象リスト!M2:M331, O2)"
        ws["P" + str(num)] = "=COUNTIFS(リンク様_SIM不良対象リスト!K2:K331, A" + str(num) + ",リンク様_SIM不良対象リスト!M2:M331, P2)"

def report():
    wb = openpyxl.load_workbook("SIM不良対象リスト_202510_サマリ1_第二集計.xlsx")
    ws = wb['リンク様_SIM不良対象リスト']  # ワークシートを指定
    ws = wb.active  # アクティブなワークシートを選択
    
    do_loop(ws)

    wb.save('SIM不良対象リスト_202510_サマリ1_第三集計.xlsx')

if __name__ == '__main__':
    report()
