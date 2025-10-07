# -*- coding: utf-8 -*-

# エビデンスの画像をエクセル形式でまとめる

import openpyxl
from openpyxl import Workbook

def do_loop(ws):

    #num = 5
    for num in range(5, 332):
        ws["J" + str(num)] = "=IF(COUNTIF(セットアップリスト!R3:R1540,0&D" + str(num) + ")>0,1,0)"
        ws["K" + str(num)] = "=XLOOKUP(0&D" + str(num) + ", セットアップリスト!R3:R1540, セットアップリスト!L3:L1540, \"見つかりません\", 0)"
        ws["L" + str(num)] = "=XLOOKUP(0&D" + str(num) + ", セットアップリスト!R3:R1540, セットアップリスト!T3:T1540, \"見つかりません\", 0)"

def report():
    wb = openpyxl.load_workbook("SIM不良対象リスト_202510.xlsx")
    ws = wb['リンク様_SIM不良対象リスト']  # ワークシートを指定
    ws = wb.active  # アクティブなワークシートを選択
    
    do_loop(ws)

    wb.save('SIM不良対象リスト_202510_第一集計.xlsx')

if __name__ == '__main__':
    report()
