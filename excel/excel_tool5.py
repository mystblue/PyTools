# -*- coding: utf-8 -*-

# エビデンスの画像をエクセル形式でまとめる

import openpyxl
from openpyxl import Workbook

def do_loop(ws):

    #ベルーナ
    for num in range(4, 36):
        ws["D" + str(num)] = "=COUNTIFS(セットアップリスト!L2:L1540, A2, セットアップリスト!T2:T1540, B" + str(num)  + ", セットアップリスト!U2:U1540, C" + str(num) + ")"
        ws["E" + str(num)] = "=COUNTIFS(リンク様_SIM不良対象リスト!K2:K331, A2, リンク様_SIM不良対象リスト!L2:L331, B" + str(num) + ", リンク様_SIM不良対象リスト!N2:N331, C" + str(num) + ")"

def report():
    wb = openpyxl.load_workbook("SIM不良対象リスト_202510_サマリ2_第４集計.xlsx")
    ws = wb['サマリ2']  # ワークシートを指定
    ws = wb.active  # アクティブなワークシートを選択
    
    do_loop(ws)

    wb.save('SIM不良対象リスト_202510_FIX.xlsx')

if __name__ == '__main__':
    report()
