# -*- coding: utf-8 -*-

# エビデンスの画像をエクセル形式でまとめる

import openpyxl
from openpyxl import Workbook

def do_loop(ws):

    #ベルーナ
    for num in range(3, 36):
        ws["E" + str(num)] = "=COUNTIFS(セットアップリスト!L2:L1540, A2, セットアップリスト!T2:T1540, B" + str(num) + ", セットアップリスト!U2:U1540, C" + str(num) + ", セットアップリスト!V2:V1540, D" + str(num) + ")"
        ws["F" + str(num)] = "=COUNTIFS(リンク様_SIM不良対象リスト!K2:K331, A2, リンク様_SIM不良対象リスト!L2:L331, B" + str(num) + ", リンク様_SIM不良対象リスト!N2:N331, C" + str(num) + ", リンク様_SIM不良対象リスト!O2:O331, D" + str(num) + ")"

    #サードウェーブ
    for num in range(36, 95):
        ws["E" + str(num)] = "=COUNTIFS(セットアップリスト!L2:L1540, A36, セットアップリスト!T2:T1540, B" + str(num) + ", セットアップリスト!U2:U1540, C" + str(num) + ", セットアップリスト!V2:V1540, D" + str(num) + ")"
        ws["F" + str(num)] = "=COUNTIFS(リンク様_SIM不良対象リスト!K2:K331, A36, リンク様_SIM不良対象リスト!L2:L331, B" + str(num) + ", リンク様_SIM不良対象リスト!N2:N331, C" + str(num) + ", リンク様_SIM不良対象リスト!O2:O331, D" + str(num) + ")"

    #山忠
    for num in range(95, 97):
        ws["E" + str(num)] = "=COUNTIFS(セットアップリスト!L2:L1540, A95, セットアップリスト!T2:T1540, B" + str(num) + ", セットアップリスト!U2:U1540, C" + str(num) + ", セットアップリスト!V2:V1540, D" + str(num) + ")"
        ws["F" + str(num)] = "=COUNTIFS(リンク様_SIM不良対象リスト!K2:K331, A95, リンク様_SIM不良対象リスト!L2:L331, B" + str(num) + ", リンク様_SIM不良対象リスト!N2:N331, C" + str(num) + ", リンク様_SIM不良対象リスト!O2:O331, D" + str(num) + ")"

    #天満屋
    for num in range(97, 109):
        ws["E" + str(num)] = "=COUNTIFS(セットアップリスト!L2:L1540, A97, セットアップリスト!T2:T1540, B" + str(num) + ", セットアップリスト!U2:U1540, C" + str(num) + ", セットアップリスト!V2:V1540, D" + str(num) + ")"
        ws["F" + str(num)] = "=COUNTIFS(リンク様_SIM不良対象リスト!K2:K331, A97, リンク様_SIM不良対象リスト!L2:L331, B" + str(num) + ", リンク様_SIM不良対象リスト!N2:N331, C" + str(num) + ", リンク様_SIM不良対象リスト!O2:O331, D" + str(num) + ")"

    #新中央航空
    for num in range(109, 113):
        ws["E" + str(num)] = "=COUNTIFS(セットアップリスト!L2:L1540, A109, セットアップリスト!T2:T1540, B" + str(num) + ", セットアップリスト!U2:U1540, C" + str(num) + ", セットアップリスト!V2:V1540, D" + str(num) + ")"
        ws["F" + str(num)] = "=COUNTIFS(リンク様_SIM不良対象リスト!K2:K331, A109, リンク様_SIM不良対象リスト!L2:L331, B" + str(num) + ", リンク様_SIM不良対象リスト!N2:N331, C" + str(num) + ", リンク様_SIM不良対象リスト!O2:O331, D" + str(num) + ")"

    #サンスター
    for num in range(113, 118):
        ws["E" + str(num)] = "=COUNTIFS(セットアップリスト!L2:L1540, A113, セットアップリスト!T2:T1540, B" + str(num) + ", セットアップリスト!U2:U1540, C" + str(num) + ", セットアップリスト!V2:V1540, D" + str(num) + ")"
        ws["F" + str(num)] = "=COUNTIFS(リンク様_SIM不良対象リスト!K2:K331, A113, リンク様_SIM不良対象リスト!L2:L331, B" + str(num) + ", リンク様_SIM不良対象リスト!N2:N331, C" + str(num) + ", リンク様_SIM不良対象リスト!O2:O331, D" + str(num) + ")"

    #ぴあ
    for num in range(118, 126):
        ws["E" + str(num)] = "=COUNTIFS(セットアップリスト!L2:L1540, A118, セットアップリスト!T2:T1540, B" + str(num) + ", セットアップリスト!U2:U1540, C" + str(num) + ", セットアップリスト!V2:V1540, D" + str(num) + ")"
        ws["F" + str(num)] = "=COUNTIFS(リンク様_SIM不良対象リスト!K2:K331, A118, リンク様_SIM不良対象リスト!L2:L331, B" + str(num) + ", リンク様_SIM不良対象リスト!N2:N331, C" + str(num) + ", リンク様_SIM不良対象リスト!O2:O331, D" + str(num) + ")"

    #小豆島
    for num in range(126, 128):
        ws["E" + str(num)] = "=COUNTIFS(セットアップリスト!L2:L1540, A126, セットアップリスト!T2:T1540, B" + str(num) + ", セットアップリスト!U2:U1540, C" + str(num) + ", セットアップリスト!V2:V1540, D" + str(num) + ")"
        ws["F" + str(num)] = "=COUNTIFS(リンク様_SIM不良対象リスト!K2:K331, A126, リンク様_SIM不良対象リスト!L2:L331, B" + str(num) + ", リンク様_SIM不良対象リスト!N2:N331, C" + str(num) + ", リンク様_SIM不良対象リスト!O2:O331, D" + str(num) + ")"

    #ピュール
    for num in range(128, 140):
        ws["E" + str(num)] = "=COUNTIFS(セットアップリスト!L2:L1540, A128, セットアップリスト!T2:T1540, B" + str(num) + ", セットアップリスト!U2:U1540, C" + str(num) + ", セットアップリスト!V2:V1540, D" + str(num) + ")"
        ws["F" + str(num)] = "=COUNTIFS(リンク様_SIM不良対象リスト!K2:K331, A128, リンク様_SIM不良対象リスト!L2:L331, B" + str(num) + ", リンク様_SIM不良対象リスト!N2:N331, C" + str(num) + ", リンク様_SIM不良対象リスト!O2:O331, D" + str(num) + ")"

    #ユーグレナ
    for num in range(140, 147):
        ws["E" + str(num)] = "=COUNTIFS(セットアップリスト!L2:L1540, A140, セットアップリスト!T2:T1540, B" + str(num) + ", セットアップリスト!U2:U1540, C" + str(num) + ", セットアップリスト!V2:V1540, D" + str(num) + ")"
        ws["F" + str(num)] = "=COUNTIFS(リンク様_SIM不良対象リスト!K2:K331, A140, リンク様_SIM不良対象リスト!L2:L331, B" + str(num) + ", リンク様_SIM不良対象リスト!N2:N331, C" + str(num) + ", リンク様_SIM不良対象リスト!O2:O331, D" + str(num) + ")"

    #水晶院
    for num in range(147, 150):
        ws["E" + str(num)] = "=COUNTIFS(セットアップリスト!L2:L1540, A147, セットアップリスト!T2:T1540, B" + str(num) + ", セットアップリスト!U2:U1540, C" + str(num) + ", セットアップリスト!V2:V1540, D" + str(num) + ")"
        ws["F" + str(num)] = "=COUNTIFS(リンク様_SIM不良対象リスト!K2:K331, A147, リンク様_SIM不良対象リスト!L2:L331, B" + str(num) + ", リンク様_SIM不良対象リスト!N2:N331, C" + str(num) + ", リンク様_SIM不良対象リスト!O2:O331, D" + str(num) + ")"

def report():
    wb = openpyxl.load_workbook("SIM不良対象リスト_202510_サマリ2_第５集計.xlsx")
    ws = wb['サマリ2']  # ワークシートを指定
    ws = wb.active  # アクティブなワークシートを選択
    
    do_loop(ws)

    wb.save('SIM不良対象リスト_202510_FIX.xlsx')

if __name__ == '__main__':
    report()
